#!/usr/bin/env python3
"""
Multi-Factor Daily Scalping Opportunity Scanner + Backtester.

Reads OHLCV, Trades, Quotes, Options Features, and Option Chains.
Ranks stocks daily using look-ahead-safe multi-factor scoring,
confirms signals across independent data sources, adaptively selects
strategies, and backtests with realistic execution.

Usage:
    source venv/bin/activate
    python scripts/strategy/scalping/scalping_analysis.py --year 2025 --top_n 40 --num_trades 50 --risk-amount 1 --strategies 'VWAP Reversion' --rr 2.0 --output data/scalping_multi_2025.xlsx

Trade-window control (optional):
    --num_trades N           cap trades backtested per strategy×R:R (default: all).
    --date_range YYYYMMDD-YYYYMMDD   restrict the analysis to a date window.
    If neither is given, the whole --year is used.

Position sizing note:
    --risk-amount N  means "risk N% of starting-cash per trade" (actual dollars at risk).
    margin_used is NOT the same as risk — margin is the buying-power requirement
    (= position_value * margin_rate). Use --max-position-pct to cap position size
    as % of equity if you want to limit margin consumption.
"""

import argparse
import datetime
import logging
import os
import sys
import time
from collections import defaultdict
from concurrent.futures import ProcessPoolExecutor, as_completed
from pathlib import Path

import numpy as np
import pandas as pd

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ── Constants ────────────────────────────────────────────────────────

RR_RATIOS = [1.5, 2.0, 2.5]

TIME_ZONES = [
    ("09:30-09:45", (9, 30), (9, 45)),
    ("09:45-10:00", (9, 45), (10, 0)),
    ("10:00-10:30", (10, 0), (10, 30)),
    ("10:30-11:30", (10, 30), (11, 30)),
    ("11:30-13:00", (11, 30), (13, 0)),
    ("13:00-14:30", (13, 0), (14, 30)),
    ("14:30-15:00", (14, 30), (15, 0)),
    ("15:00-16:00", (15, 0), (16, 0)),
]

DATA_OHLCV = Path("data/SPY")
DATA_TRADES = Path("data/trades")
DATA_QUOTES = Path("data/quotes")
DATA_OPTIONS = Path("data/options/stocks")
DATA_CHAINS = Path("data/options/chains")

DEFAULT_AGGREGATE = "1sec"

ALL_STRATEGIES = [
    "Momentum Breakout",
    "VWAP Reversion",
    "RSI Scalp",
    "EMA Pullback",
    "Opening Range Breakout",
    "Delta Confirmation Breakout",
    "Quote Imbalance Breakout",
    "VWAP + Delta",
    "Liquidity Vacuum / Exhaustion Fade",
    "VWAP + Momentum Confluence",
    "VWAP + Quote Imbalance Confluence",
    "VWAP + Delta Confluence",
    "VWAP Triple Confluence",
    "RSI + Momentum Confluence",
    "RSI + Delta Confluence",
    "EMA + Momentum Confluence",
    "EMA + Delta Confluence",
    "ORB + Volume Confluence",
    "Momentum + Quote Confluence",
    "Momentum + Delta + Quote Confluence",
]


# ══════════════════════════════════════════════════════════════════════
# SECTION 1: Data Availability Report & Loaders
# ══════════════════════════════════════════════════════════════════════

def _path(ticker: str, base: Path, year: str, agg: str, kind: str = "") -> Path:
    name = f"{ticker}_{year}_{agg}_{kind}.csv" if kind else f"{ticker}_{year}_{agg}.csv"
    return base / agg / year / name


def load_ohlcv(ticker: str, year: str, agg: str = DEFAULT_AGGREGATE) -> pd.DataFrame | None:
    p = _path(ticker, DATA_OHLCV, year, agg)
    if not p.exists():
        return None
    df = pd.read_csv(p)
    df["timestamp"] = pd.to_datetime(df["timestamp"], utc=True)
    df.set_index("timestamp", inplace=True)
    df.sort_index(inplace=True)
    return df


def load_trades(ticker: str, year: str, agg: str = DEFAULT_AGGREGATE) -> pd.DataFrame | None:
    p = _path(ticker, DATA_TRADES, year, agg, "trades")
    if not p.exists():
        return None
    df = pd.read_csv(p)
    df["timestamp"] = pd.to_datetime(df["timestamp"], utc=True)
    df.set_index("timestamp", inplace=True)
    df.sort_index(inplace=True)
    return df


def load_quotes(ticker: str, year: str, agg: str = DEFAULT_AGGREGATE) -> pd.DataFrame | None:
    p = _path(ticker, DATA_QUOTES, year, agg, "quotes")
    if not p.exists():
        return None
    df = pd.read_csv(p)
    df["timestamp"] = pd.to_datetime(df["timestamp"], utc=True)
    df.set_index("timestamp", inplace=True)
    df.sort_index(inplace=True)
    return df


def load_options_features(ticker: str, year: str, agg: str = DEFAULT_AGGREGATE) -> pd.DataFrame | None:
    p = _path(ticker, DATA_OPTIONS, year, agg, "options")
    if not p.exists():
        return None
    df = pd.read_csv(p)
    df["timestamp"] = pd.to_datetime(df["timestamp"], utc=True)
    df.set_index("timestamp", inplace=True)
    df.sort_index(inplace=True)
    return df


def load_chains(ticker: str, year: str, agg: str = DEFAULT_AGGREGATE) -> pd.DataFrame | None:
    """Load option chain data. Prefer `agg`, fall back to 1min."""
    aggs = [agg] if agg == "1min" else [agg, "1min"]
    for a in aggs:
        p = _path(ticker, DATA_CHAINS, year, a, "chains")
        if p.exists():
            df = pd.read_csv(p)
            df["timestamp"] = pd.to_datetime(df["timestamp"], utc=True)
            return df
    return None


def report_data_availability(year: str, agg: str = DEFAULT_AGGREGATE) -> tuple[dict, dict]:
    """Scan all data directories. Returns (per_ticker_map, counts)."""
    ohlcv_dir = DATA_OHLCV / agg / year
    tickers_all = sorted(
        [f.stem.split("_")[0] for f in ohlcv_dir.glob(f"*_{year}_{agg}.csv")]
    ) if ohlcv_dir.exists() else []

    report = {}
    for t in tickers_all:
        report[t] = {
            "OHLCV": _path(t, DATA_OHLCV, year, agg).exists(),
            "Trades": _path(t, DATA_TRADES, year, agg, "trades").exists(),
            "Quotes": _path(t, DATA_QUOTES, year, agg, "quotes").exists(),
            "Options": _path(t, DATA_OPTIONS, year, agg, "options").exists(),
            "Chains": (
                _path(t, DATA_CHAINS, year, agg, "chains").exists()
                or (agg != "1min" and _path(t, DATA_CHAINS, year, "1min", "chains").exists())
            ),
        }

    counts = {}
    for col in ["OHLCV", "Trades", "Quotes", "Options", "Chains"]:
        counts[col] = sum(1 for v in report.values() if v.get(col))

    return report, counts


# ══════════════════════════════════════════════════════════════════════
# SECTION 2: Capital Manager
# ══════════════════════════════════════════════════════════════════════

class CapitalManager:
    """Tracks cash, equity, margin, and position sizing for a scalping account.

    Key distinction (read carefully):
      - risk_$       = dollars actually at risk per trade (= risk_pct% of starting_cash)
      - margin_used  = buying-power requirement to hold the position
                        (= position_value * margin_rate)
      - position_value = shares * entry_price

    The user's risk is capped at risk_$ via the stop distance. margin_used can be
    much larger than risk_$ when stops are tight relative to entry price. That is
    normal margin-account behaviour — risk_$ is what matters for loss sizing.
    Use --max-position-pct to cap position_value as % of equity if you want to
    limit concurrent margin use.
    """

    def __init__(
        self,
        starting_cash: float = 100_000.0,
        risk_pct: float = 0.10,
        compounding: bool = False,
        margin_rate: float = 0.25,
        max_position_pct: float = 2.0,
    ):
        self.starting_cash = starting_cash
        self.risk_pct = risk_pct          # fraction of starting_cash risked per trade
        self.compounding = compounding
        self.margin_rate = margin_rate
        self.leverage = 1.0 / margin_rate
        self.max_position_pct = max_position_pct  # fraction of equity

        self.cash = starting_cash
        self.equity = starting_cash
        self.peak_equity = starting_cash
        self.equity_curve: list[dict] = []
        self.trade_count = 0
        self.total_commission = 0.0

    def reset(self):
        self.cash = self.starting_cash
        self.equity = self.starting_cash
        self.peak_equity = self.starting_cash
        self.equity_curve = []
        self.trade_count = 0
        self.total_commission = 0.0

    def _risk_base(self) -> float:
        return self.equity if self.compounding else self.starting_cash

    def _risk_dollars(self) -> float:
        return self._risk_base() * self.risk_pct

    def calculate_position(self, entry: float, stop: float, direction: int):
        """Returns (shares, position_value, margin_required, risk_dollars) or None."""
        sd = abs(entry - stop)
        if sd <= 0:
            return None
        risk_d = self._risk_dollars()
        raw_shares = risk_d / sd
        pv = raw_shares * entry

        if pv > self.equity * self.leverage:
            return None
        if pv > self.equity * self.max_position_pct:
            return None
        shares = int(raw_shares)
        if shares <= 0:
            return None
        return (shares, shares * entry, shares * entry * self.margin_rate, risk_d)

    def execute_trade(self, entry: float, exit: float, stop: float, direction: int, ts=None):
        pos = self.calculate_position(entry, stop, direction)
        if pos is None:
            return None
        shares, pv, margin, risk_d = pos
        if margin > self.cash:
            return None
        self.cash -= margin
        gross = (exit - entry) * shares * direction
        comm = max(0.35, shares * 0.0001)
        net = gross - comm
        self.cash += margin + net
        self.equity = self.cash
        self.peak_equity = max(self.peak_equity, self.equity)
        self.total_commission += comm
        self.trade_count += 1
        dd = (self.peak_equity - self.equity) / self.peak_equity * 100 if self.peak_equity > 0 else 0.0
        self.equity_curve.append({"timestamp": ts, "equity": self.equity, "drawdown_pct": dd})
        return {
            "shares": shares,
            "position_value": round(pv, 2),
            "margin_used": round(margin, 2),
            "risk_$": round(risk_d, 2),
            "pnl_$": round(net, 2),
            "pnl_%_capital": round(net / self.starting_cash * 100, 4),
            "commission_$": round(comm, 2),
            "running_equity": round(self.equity, 2),
            "drawdown_%": round(dd, 2),
        }

    def total_return_pct(self) -> float:
        return (self.equity - self.starting_cash) / self.starting_cash * 100


# ══════════════════════════════════════════════════════════════════════
# SECTION 3: Daily Feature Computation (Look-Ahead Safe)
# ══════════════════════════════════════════════════════════════════════

def atr_arr(high, low, close, period=14):
    tr = np.maximum(
        high - low,
        np.maximum(np.abs(high - np.roll(close, 1)), np.abs(low - np.roll(close, 1))),
    )
    tr[0] = high[0] - low[0]
    return pd.Series(tr).rolling(period).mean().values


def rsi_arr(close, period=14):
    delta = np.diff(close, prepend=close[0])
    gain = np.maximum(delta, 0)
    loss = np.maximum(-delta, 0)
    avg_gain = pd.Series(gain).rolling(period).mean().values
    avg_loss = pd.Series(loss).rolling(period).mean().values
    rs = np.divide(avg_gain, avg_loss, out=np.zeros_like(avg_gain), where=avg_loss != 0)
    return 100 - 100 / (1 + rs)


def _parse_date_range(date_range: str | None) -> tuple | None:
    """Parse 'YYYYMMDD-YYYYMMDD' into (start_date, end_date) datetime.date tuple."""
    if not date_range:
        return None
    parts = date_range.split("-")
    if len(parts) != 2:
        raise SystemExit("Error: --date_range must be YYYYMMDD-YYYYMMDD (e.g. 20250101-20250331)")
    try:
        start = datetime.datetime.strptime(parts[0].strip(), "%Y%m%d").date()
        end = datetime.datetime.strptime(parts[1].strip(), "%Y%m%d").date()
    except ValueError:
        raise SystemExit("Error: --date_range must be YYYYMMDD-YYYYMMDD (e.g. 20250101-20250331)")
    if end < start:
        raise SystemExit("Error: --date_range end is before start")
    return (start, end)


def _filter_date_range(df: pd.DataFrame | None, date_range: tuple | None) -> pd.DataFrame | None:
    """Restrict a DatetimeIndex-indexed frame to [start, end] (inclusive, UTC dates)."""
    if df is None or date_range is None:
        return df
    start_d, end_d = date_range
    start_ts = pd.Timestamp(start_d, tz="UTC")
    end_ts = pd.Timestamp(end_d, tz="UTC") + pd.Timedelta(days=1)
    return df[(df.index >= start_ts) & (df.index < end_ts)]


def compute_daily_features(ticker: str, year: str, agg: str = DEFAULT_AGGREGATE,
                           date_range: tuple | None = None) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Load all datasets, align, and compute daily opening features.
    Returns (aligned_minute_df, daily_rank_df)."""

    ohlcv = load_ohlcv(ticker, year, agg)
    ohlcv = _filter_date_range(ohlcv, date_range)
    if ohlcv is None or len(ohlcv) < 500:
        return pd.DataFrame(), pd.DataFrame()

    trades_df = _filter_date_range(load_trades(ticker, year, agg), date_range)
    quotes = _filter_date_range(load_quotes(ticker, year, agg), date_range)
    options = _filter_date_range(load_options_features(ticker, year, agg), date_range)
    chains = load_chains(ticker, year, agg)

    # ── Build aligned 1-min dataframe ──
    df = ohlcv[["open", "high", "low", "close", "volume"]].copy()
    df.columns = ["o", "h", "l", "c", "vol"]

    # Merge trades
    if trades_df is not None:
        t_cols = [c for c in ["trade_count", "volume", "buy_volume", "sell_volume",
                               "delta", "cumulative_delta", "delta_pct",
                               "aggression_ratio", "trade_frequency",
                               "large_trade_count", "large_trade_ratio",
                               "avg_seconds_between_trades",
                               "avg_trade_size", "largest_trade", "stddev_trade_size"]
                  if c in trades_df.columns]
        if t_cols:
            df = df.join(trades_df[t_cols].add_prefix("tr_"), how="left")

    # Merge quotes
    if quotes is not None:
        q_cols = [c for c in ["avg_spread", "max_spread", "min_spread",
                               "avg_bid", "avg_ask", "quote_imbalance",
                               "bid_size", "ask_size", "spread_volatility",
                               "mid_price", "microprice"]
                  if c in quotes.columns]
        if q_cols:
            df = df.join(quotes[q_cols].add_prefix("qu_"), how="left")

    # Merge options features
    if options is not None:
        o_cols = [c for c in ["underlying_price", "atm_strike", "atm_call_close",
                               "atm_put_close", "atm_straddle_price", "expected_move",
                               "put_volume", "call_volume", "put_call_ratio",
                               "contract_count", "avg_bid_ask_spread",
                               "avg_bid_size", "avg_ask_size", "quote_imbalance",
                               "atm_days_to_expiry", "iv30d_call_close", "iv30d_put_close"]
                  if c in options.columns]
        if o_cols:
            df = df.join(options[o_cols].add_prefix("opt_"), how="left")

    # Rolling features (vectorized, look-ahead safe)
    df["returns_1m"] = df["c"].pct_change()
    df["log_ret"] = np.log(df["c"] / df["c"].shift(1))
    df["range_1m"] = df["h"] - df["l"]
    df["range_pct"] = df["range_1m"] / df["c"] * 100

    df["atr_14"] = atr_arr(df["h"].values, df["l"].values, df["c"].values, 14)
    df["atr_pct"] = df["atr_14"] / df["c"] * 100
    df["vol_ma_20"] = df["vol"].rolling(20).mean()
    df["rel_vol"] = df["vol"] / df["vol_ma_20"].replace(0, np.nan)
    df["returns_std_20"] = df["returns_1m"].rolling(20).std()
    df["rsi_14"] = rsi_arr(df["c"].values, 14)

    # VWAP (cumulative from daily open)
    df["date"] = df.index.normalize()
    for date, grp in df.groupby("date"):
        cum_pv = (grp["c"] * grp["vol"]).cumsum()
        cum_vol = grp["vol"].cumsum()
        df.loc[grp.index, "vwap"] = cum_pv / cum_vol.replace(0, np.nan)
    df["vwap_dist_pct"] = (df["c"] - df["vwap"]) / df["vwap"] * 100

    # ── Build daily ranking rows ──
    daily_rows = []
    for date, grp in df.groupby("date"):
        if len(grp) < 10:
            continue
        row = {"ticker": ticker, "date": date}

        # --- Liquidity ---
        row["dollar_vol"] = (grp["c"] * grp["vol"]).mean()
        row["trade_freq"] = grp.get("tr_trade_frequency", pd.Series([0])).mean()
        row["avg_spread_qu"] = grp.get("qu_avg_spread", pd.Series([np.nan])).mean()
        row["spread_vol"] = grp.get("qu_spread_volatility", pd.Series([np.nan])).mean()
        row["bid_depth"] = grp.get("qu_bid_size", pd.Series([0])).mean()
        row["ask_depth"] = grp.get("qu_ask_size", pd.Series([0])).mean()

        # --- Order Flow ---
        row["tr_delta_mean"] = grp.get("tr_delta", pd.Series([0])).mean()
        row["tr_delta_sum"] = grp.get("tr_delta", pd.Series([0])).sum()
        row["tr_delta_pct_mean"] = grp.get("tr_delta_pct", pd.Series([0])).mean()
        row["tr_aggression_ratio"] = grp.get("tr_aggression_ratio", pd.Series([0])).mean()
        row["tr_large_trade_ratio"] = grp.get("tr_large_trade_ratio", pd.Series([0])).mean()
        row["tr_large_trade_count"] = grp.get("tr_large_trade_count", pd.Series([0])).sum()
        row["tr_buy_vol"] = grp.get("tr_buy_volume", pd.Series([0])).sum()
        row["tr_sell_vol"] = grp.get("tr_sell_volume", pd.Series([0])).sum()
        row["buy_sell_ratio"] = (row["tr_buy_vol"] / row["tr_sell_vol"]
                                 if row["tr_sell_vol"] > 0 else 0)

        # --- Price / Volatility ---
        row["atr_14_pct"] = grp["atr_pct"].mean()
        row["rel_vol_mean"] = grp["rel_vol"].mean()
        row["returns_std"] = grp["returns_1m"].std()
        row["range_mean_pct"] = grp["range_pct"].mean()
        row["open_price"] = grp["o"].iloc[0]
        row["close_prev_day"] = grp["c"].iloc[-1]
        row["first_30_range"] = (grp["h"].iloc[:30].max() - grp["l"].iloc[:30].min()
                                 if len(grp) >= 30 else grp["range_1m"].iloc[:len(grp)].sum())

        # --- Quotes ---
        row["qu_imbalance_mean"] = grp.get("qu_quote_imbalance", pd.Series([0])).mean()
        row["micro_mid_diff"] = ((grp.get("qu_microprice", grp["c"]) -
                                   grp.get("qu_mid_price", grp["c"])) / grp["c"] * 100).mean()
        row["spread_contraction"] = (grp.get("qu_avg_spread", pd.Series([np.nan])).iloc[:30].mean() -
                                     grp.get("qu_avg_spread", pd.Series([np.nan])).iloc[-30:].mean()
                                     if len(grp) >= 60 else 0)

        # --- Options ---
        row["opt_put_call_ratio"] = grp.get("opt_put_call_ratio", pd.Series([1])).mean()
        row["opt_straddle_pct"] = (grp.get("opt_atm_straddle_price", pd.Series([0])) / grp["c"] * 100).mean()
        row["opt_expected_move_pct"] = (grp.get("opt_expected_move", pd.Series([0])) / grp["c"] * 100).mean()
        row["opt_call_vol"] = grp.get("opt_call_volume", pd.Series([0])).sum()
        row["opt_put_vol"] = grp.get("opt_put_volume", pd.Series([0])).sum()
        row["opt_total_vol"] = row["opt_call_vol"] + row["opt_put_vol"]
        row["opt_contract_count"] = grp.get("opt_contract_count", pd.Series([0])).mean()
        row["opt_bid_ask_spread"] = grp.get("opt_avg_bid_ask_spread", pd.Series([0])).mean()

        # --- Chain features: ATM IV, skew, gamma ---
        if chains is not None and len(chains) > 0:
            chain_day = chains[chains["timestamp"].dt.normalize() == date]
            if len(chain_day) > 0:
                # ATM IV: average implied_volatility across all contracts
                if "implied_volatility" in chain_day.columns:
                    iv_vals = pd.to_numeric(chain_day["implied_volatility"], errors="coerce")
                    row["chain_atm_iv"] = iv_vals.mean()
                # Gamma exposure: sum of |gamma| * volume
                if "gamma" in chain_day.columns and "volume" in chain_day.columns:
                    gam = pd.to_numeric(chain_day["gamma"], errors="coerce").abs()
                    vol = pd.to_numeric(chain_day["volume"], errors="coerce").fillna(0)
                    row["chain_gamma_exp"] = (gam * vol).sum()
                # Put/call IV skew
                if "call_put" in chain_day.columns and "implied_volatility" in chain_day.columns:
                    calls = chain_day[chain_day["call_put"] == "C"]
                    puts = chain_day[chain_day["call_put"] == "P"]
                    call_iv = pd.to_numeric(calls["implied_volatility"], errors="coerce").mean() if len(calls) > 0 else 0
                    put_iv = pd.to_numeric(puts["implied_volatility"], errors="coerce").mean() if len(puts) > 0 else 0
                    row["chain_iv_skew"] = put_iv - call_iv
                # Delta exposure
                if "delta" in chain_day.columns:
                    row["chain_delta_sum"] = pd.to_numeric(chain_day["delta"], errors="coerce").sum()
                # Vega-weighted IV
                if "vega" in chain_day.columns and "implied_volatility" in chain_day.columns:
                    vega = pd.to_numeric(chain_day["vega"], errors="coerce").abs()
                    iv = pd.to_numeric(chain_day["implied_volatility"], errors="coerce")
                    vega_sum = vega.sum()
                    if vega_sum > 0:
                        row["chain_vega_weighted_iv"] = (iv * vega).sum() / vega_sum

        daily_rows.append(row)

    daily_df = pd.DataFrame(daily_rows)
    return df, daily_df


# ══════════════════════════════════════════════════════════════════════
# SECTION 4: Multi-Factor Scoring
# ══════════════════════════════════════════════════════════════════════

def _z(series: pd.Series) -> pd.Series:
    """Z-score clip [-3,3] → map to 0-100."""
    s = series.replace([np.inf, -np.inf], np.nan).fillna(0)
    m, std = s.mean(), s.std()
    if std == 0 or pd.isna(std):
        return pd.Series(50.0, index=s.index)
    return ((s - m) / std).clip(-3, 3).add(3).div(6).mul(100)


def score_liquidity(daily_df: pd.DataFrame) -> pd.Series:
    return (
        _z(daily_df["dollar_vol"]) * 0.25 +
        _z(daily_df["trade_freq"]) * 0.25 +
        (100 - _z(daily_df["avg_spread_qu"])) * 0.25 +
        _z(daily_df["bid_depth"] + daily_df["ask_depth"]) * 0.25
    )


def score_orderflow(daily_df: pd.DataFrame) -> pd.Series:
    return (
        _z(daily_df["tr_delta_sum"]) * 0.20 +
        _z(daily_df["tr_aggression_ratio"]) * 0.20 +
        _z(daily_df["buy_sell_ratio"]) * 0.20 +
        _z(daily_df["tr_large_trade_ratio"]) * 0.20 +
        _z(daily_df["tr_large_trade_count"]) * 0.20
    )


def score_volatility(daily_df: pd.DataFrame) -> pd.Series:
    return (
        _z(daily_df["atr_14_pct"]) * 0.30 +
        _z(daily_df["rel_vol_mean"]) * 0.25 +
        _z(daily_df["returns_std"]) * 0.20 +
        _z(daily_df["first_30_range"]) * 0.25
    )


def score_quotes(daily_df: pd.DataFrame) -> pd.Series:
    return (
        _z(daily_df["qu_imbalance_mean"].abs()) * 0.30 +
        _z(daily_df["micro_mid_diff"].abs()) * 0.25 +
        _z(daily_df["spread_contraction"]) * 0.25 +
        (100 - _z(daily_df["avg_spread_qu"])) * 0.20
    )


def score_options(daily_df: pd.DataFrame) -> pd.Series:
    pcr_signal = (daily_df["opt_put_call_ratio"] - 1).abs()
    s = (
        _z(daily_df["opt_total_vol"]) * 0.15 +
        _z(daily_df["opt_expected_move_pct"]) * 0.15 +
        _z(pcr_signal) * 0.10 +
        _z(daily_df["opt_straddle_pct"]) * 0.15 +
        _z(daily_df["chain_iv_skew"].abs()) * 0.15 +
        _z(daily_df["chain_gamma_exp"].abs()) * 0.15
    )
    # If chain_atm_iv exists, add IV expansion component
    if "chain_atm_iv" in daily_df.columns and daily_df["chain_atm_iv"].notna().any():
        s += _z(daily_df["chain_atm_iv"]) * 0.10
    if "chain_vega_weighted_iv" in daily_df.columns and daily_df["chain_vega_weighted_iv"].notna().any():
        s += _z(daily_df["chain_vega_weighted_iv"]) * 0.05
    return s


def score_cross_market(daily_df: pd.DataFrame) -> pd.Series:
    """Check for confirmation across independent data sources. 0-100."""
    c1 = ((daily_df["returns_std"] > 0.001) & (daily_df["tr_delta_sum"] > 0)).astype(float)
    c2 = ((daily_df["returns_std"] > 0.001) & (daily_df["qu_imbalance_mean"].abs() > 0)).astype(float)
    c3 = ((daily_df["tr_aggression_ratio"] > 1) &
          (daily_df["opt_call_vol"] > daily_df["opt_put_vol"] * 1.2)).astype(float)
    c4 = ((daily_df["returns_std"] > 0.001) & (daily_df["opt_straddle_pct"] > 1)).astype(float)
    c5 = ((daily_df["buy_sell_ratio"] > 1.2) & (daily_df["qu_imbalance_mean"] > 0)).astype(float)
    return (c1 + c2 + c3 + c4 + c5) / 5 * 100


def compute_opportunity_score(daily_df: pd.DataFrame) -> pd.DataFrame:
    """Add component scores and overall opportunity_score to daily ranking df."""
    if daily_df.empty or len(daily_df) < 2:
        daily_df["opportunity_score"] = 50.0
        return daily_df

    for col in ["dollar_vol", "trade_freq", "avg_spread_qu", "bid_depth", "ask_depth",
                "tr_delta_sum", "tr_aggression_ratio", "buy_sell_ratio",
                "tr_large_trade_ratio", "tr_large_trade_count",
                "atr_14_pct", "rel_vol_mean", "returns_std", "first_30_range",
                "qu_imbalance_mean", "micro_mid_diff", "spread_contraction",
                "opt_total_vol", "opt_expected_move_pct", "opt_put_call_ratio",
                "opt_straddle_pct", "chain_iv_skew", "chain_gamma_exp",
                "chain_atm_iv", "chain_vega_weighted_iv"]:
        if col not in daily_df.columns:
            daily_df[col] = 0
        daily_df[col] = daily_df[col].fillna(0)

    scores = pd.DataFrame(index=daily_df.index)

    scores["liquidity_score"] = score_liquidity(daily_df)
    scores["orderflow_score"] = score_orderflow(daily_df)
    scores["volatility_score"] = score_volatility(daily_df)
    scores["quotes_score"] = score_quotes(daily_df)
    scores["options_score"] = score_options(daily_df)
    scores["cross_market_score"] = score_cross_market(daily_df)

    scores["opportunity_score"] = (
        scores["liquidity_score"] * 0.15 +
        scores["orderflow_score"] * 0.20 +
        scores["volatility_score"] * 0.20 +
        scores["quotes_score"] * 0.15 +
        scores["options_score"] * 0.15 +
        scores["cross_market_score"] * 0.15
    )

    scores = scores.round(1)
    return pd.concat([daily_df, scores], axis=1)


# ══════════════════════════════════════════════════════════════════════
# SECTION 5: Strategy Backtests
# ══════════════════════════════════════════════════════════════════════

def _walk_trade(df, entry_i, entry, stop, target, d, max_bars):
    """Common trade-walk logic. Returns trade dict or None (no resolution within window)."""
    h, l, c, ts = df["h"].values, df["l"].values, df["c"].values, df.index
    for j in range(entry_i + 1, min(entry_i + max_bars, len(df))):
        hit_t = h[j] >= target if d > 0 else l[j] <= target
        hit_s = l[j] <= stop if d > 0 else h[j] >= stop
        if hit_t:
            return {"entry_time": ts[entry_i], "exit_time": ts[j], "direction": "long" if d > 0 else "short",
                    "entry_price": entry, "exit_price": target, "stop_price": stop,
                    "pnl_%": round((target - entry) / entry * 100 * d, 4), "result": "win", "bars_held": j - entry_i}
        elif hit_s:
            return {"entry_time": ts[entry_i], "exit_time": ts[j], "direction": "long" if d > 0 else "short",
                    "entry_price": entry, "exit_price": stop, "stop_price": stop,
                    "pnl_%": round((stop - entry) / entry * 100 * d, 4), "result": "loss", "bars_held": j - entry_i}
    j = min(entry_i + max_bars, len(df)) - 1
    return {"entry_time": ts[entry_i], "exit_time": ts[j], "direction": "long" if d > 0 else "short",
            "entry_price": entry, "exit_price": c[j], "stop_price": stop,
            "pnl_%": (c[j] - entry) / entry * 100 * d, "result": "timeout", "bars_held": j - entry_i}


def backtest_momentum_breakout(df, rr=1.5, max_bars=20):
    results = []
    n = 5
    h, l, c, v, ts = df["h"].values, df["l"].values, df["c"].values, df["vol"].values, df.index
    for i in range(n + 5, len(df)):
        rh, rl = np.max(h[i - n:i]), np.min(l[i - n:i])
        avg_v = np.mean(v[i - n:i])
        if h[i] > rh and v[i] > avg_v * 1.2:
            entry, stop = rh + 0.01, rl
            risk = entry - stop
            if risk <= 0: continue
            target = entry + risk * rr
            r = _walk_trade(df, i, entry, stop, target, 1, max_bars)
            if r: results.append(r)
    return results


def backtest_vwap_reversion(df, rr=1.5, max_bars=20, dev=0.3):
    results = []
    h, l, c, v, ts = df["h"].values, df["l"].values, df["c"].values, df["vol"].values, df.index
    period = 20
    for i in range(period + 5, len(df)):
        vwap = np.sum(c[i - period:i] * v[i - period:i]) / (np.sum(v[i - period:i]) + 1e-9)
        dev_pct = (c[i] - vwap) / vwap * 100
        atr = np.mean(np.abs(np.diff(c[i - 10:i]))) if i >= 11 else 0.03
        if atr <= 0: continue
        if dev_pct > dev:
            entry = c[i]; stop = entry + atr * rr; risk = stop - entry
            target = entry - risk; d = -1
        elif dev_pct < -dev:
            entry = c[i]; stop = entry - atr * rr; risk = entry - stop
            target = entry + risk; d = 1
        else:
            continue
        if risk <= 0: continue
        r = _walk_trade(df, i, entry, stop, target, d, max_bars)
        if r: results.append(r)
    return results


def backtest_rsi_scalp(df, rr=1.5, max_bars=15):
    results = []
    h, l, c, ts = df["h"].values, df["l"].values, df["c"].values, df.index
    rv = rsi_arr(c, 5)
    for i in range(15, len(df)):
        if pd.isna(rv[i - 1]) or pd.isna(rv[i]): continue
        atr = np.mean(np.abs(np.diff(c[i - 10:i]))) if i >= 11 else 0.03
        if atr <= 0: continue
        if rv[i - 1] < 30 and rv[i] >= 30:
            entry = c[i]; stop = entry - atr * 2; risk = entry - stop
            target = entry + risk * rr; d = 1
        elif rv[i - 1] > 70 and rv[i] <= 70:
            entry = c[i]; stop = entry + atr * 2; risk = stop - entry
            target = entry - risk * rr; d = -1
        else:
            continue
        if risk <= 0: continue
        r = _walk_trade(df, i, entry, stop, target, d, max_bars)
        if r: results.append(r)
    return results


def backtest_ema_pullback(df, rr=1.5, max_bars=15):
    results = []
    h, l, c, ts = df["h"].values, df["l"].values, df["c"].values, df.index
    ema = pd.Series(c).ewm(span=20, adjust=False).mean().values
    for i in range(40, len(df)):
        trend = ema[i] - ema[i - 5]
        atr = np.mean(np.abs(np.diff(c[i - 14:i]))) if i >= 15 else 0.03
        if atr <= 0: continue
        prox = (c[i] - ema[i]) / ema[i] * 100
        if trend > 0 and -0.2 < prox < 0.15:
            entry = c[i]; stop = np.min(l[i - 3:i]) - atr * 0.5
            risk = entry - stop; target = entry + risk * rr; d = 1
        elif trend < 0 and -0.15 < prox < 0.2:
            entry = c[i]; stop = np.max(h[i - 3:i]) + atr * 0.5
            risk = stop - entry; target = entry - risk * rr; d = -1
        else:
            continue
        if risk <= 0: continue
        r = _walk_trade(df, i, entry, stop, target, d, max_bars)
        if r: results.append(r)
    return results


def backtest_opening_range_breakout(df, rr=1.5, max_bars=30, lookback=5):
    results = []
    h, l, c, ts = df["h"].values, df["l"].values, df["c"].values, df.index
    try:
        dates = df.tz_convert("US/Eastern").index.normalize()
    except Exception:
        dates = df.index.normalize()

    n = len(dates)
    if n == 0: return results

    orb_high = orb_low = 0.0
    orb_seen = 0
    orb_ready = False
    prev_date = dates[0]

    for i in range(n):
        if i > 0 and dates[i] != prev_date:
            orb_seen = 0; orb_high = orb_low = 0.0; orb_ready = False
            prev_date = dates[i]

        if not orb_ready and orb_seen < lookback:
            orb_high = h[i] if orb_seen == 0 else max(orb_high, h[i])
            orb_low = l[i] if orb_seen == 0 else min(orb_low, l[i])
            orb_seen += 1
            if orb_seen == lookback:
                orb_ready = True
            continue

        if not orb_ready:
            continue

        try:
            cur_hour = ts[i].tz_convert("US/Eastern").hour
        except Exception:
            cur_hour = 9
        if cur_hour < 9 or cur_hour >= 11:
            continue

        if h[i] > orb_high:
            entry = orb_high + 0.01; stop = orb_low
            risk = entry - stop
            if risk <= 0: continue
            target = entry + risk * rr
            r = _walk_trade(df, i, entry, stop, target, 1, max_bars)
            if r: results.append(r)
        elif l[i] < orb_low:
            entry = orb_low - 0.01; stop = orb_high
            risk = stop - entry
            if risk <= 0: continue
            target = entry - risk * rr
            r = _walk_trade(df, i, entry, stop, target, -1, max_bars)
            if r: results.append(r)
    return results


def backtest_delta_breakout(df, rr=1.5, max_bars=15):
    results = []
    n = 5
    h, l, v, ts = df["h"].values, df["l"].values, df["vol"].values, df.index
    delta_col = "tr_delta" if "tr_delta" in df.columns else None
    for i in range(n + 5, len(df)):
        rh, rl = np.max(h[i - n:i]), np.min(l[i - n:i])
        avg_v = np.mean(v[i - n:i])
        delta_ok = True
        if delta_col:
            dv = df[delta_col].iloc[i]
            delta_ok = dv > 0 if pd.notna(dv) else True
        if h[i] > rh and v[i] > avg_v * 1.2 and delta_ok:
            entry = rh + 0.01; stop = rl
            risk = entry - stop
            if risk <= 0: continue
            target = entry + risk * rr
            r = _walk_trade(df, i, entry, stop, target, 1, max_bars)
            if r: results.append(r)
    return results


def backtest_quote_imbalance_breakout(df, rr=1.5, max_bars=15):
    results = []
    n = 5
    h, l, v, ts = df["h"].values, df["l"].values, df["vol"].values, df.index
    qi_col = "qu_quote_imbalance" if "qu_quote_imbalance" in df.columns else None
    for i in range(n + 5, len(df)):
        rh, rl = np.max(h[i - n:i]), np.min(l[i - n:i])
        avg_v = np.mean(v[i - n:i])
        qi_ok = True
        if qi_col:
            qi = df[qi_col].iloc[i]
            qi_ok = qi > 0 if pd.notna(qi) else True
        if h[i] > rh and v[i] > avg_v * 1.2 and qi_ok:
            entry = rh + 0.01; stop = rl
            risk = entry - stop
            if risk <= 0: continue
            target = entry + risk * rr
            r = _walk_trade(df, i, entry, stop, target, 1, max_bars)
            if r: results.append(r)
    return results


def backtest_vwap_delta(df, rr=1.5, max_bars=20):
    """VWAP reversion confirmed by delta direction (order-flow aware)."""
    results = []
    h, l, c, v, ts = df["h"].values, df["l"].values, df["c"].values, df["vol"].values, df.index
    period = 20
    delta_col = "tr_delta" if "tr_delta" in df.columns else None
    for i in range(period + 5, len(df)):
        vwap = np.sum(c[i - period:i] * v[i - period:i]) / (np.sum(v[i - period:i]) + 1e-9)
        dev_pct = (c[i] - vwap) / vwap * 100
        atr = np.mean(np.abs(np.diff(c[i - 10:i]))) if i >= 11 else 0.03
        if atr <= 0: continue

        # Check delta direction: below VWAP + positive delta = buy signal
        delta_val = 0
        if delta_col:
            dv = df[delta_col].iloc[i]
            delta_val = dv if pd.notna(dv) else 0

        if dev_pct > 0.3 and delta_val < 0:
            entry = c[i]; stop = entry + atr * rr; risk = stop - entry
            target = entry - risk; d = -1
        elif dev_pct < -0.3 and delta_val > 0:
            entry = c[i]; stop = entry - atr * rr; risk = entry - stop
            target = entry + risk; d = 1
        # Fall back to standard VWAP reversion if no delta data
        elif dev_pct > 0.3:
            entry = c[i]; stop = entry + atr * rr; risk = stop - entry
            target = entry - risk; d = -1
        elif dev_pct < -0.3:
            entry = c[i]; stop = entry - atr * rr; risk = entry - stop
            target = entry + risk; d = 1
        else:
            continue
        if risk <= 0: continue
        r = _walk_trade(df, i, entry, stop, target, d, max_bars)
        if r: results.append(r)
    return results


def backtest_liquidity_vacuum(df, rr=1.5, max_bars=15):
    """Liquidity Vacuum / Exhaustion Fade: fade extreme volume spikes that fail to move price."""
    results = []
    h, l, c, v, ts = df["h"].values, df["l"].values, df["c"].values, df["vol"].values, df.index
    vol_ma = pd.Series(v).rolling(50).mean().values
    for i in range(55, len(df)):
        if pd.isna(vol_ma[i]) or vol_ma[i] <= 0: continue
        vol_spike = v[i] / vol_ma[i]
        if vol_spike < 2.5: continue  # require 2.5x volume surge

        range_1m = h[i] - l[i]
        avg_range = np.mean(np.abs(h[i - 20:i] - l[i - 20:i]))
        if avg_range <= 0: continue

        # Exhaustion: high volume but small range → liquidity vacuum, price about to move
        if range_1m < avg_range * 0.7:
            atr = np.mean(np.abs(np.diff(c[i - 10:i]))) if i >= 11 else 0.03
            if atr <= 0: continue
            # Fade the direction of the previous move (counter-trend)
            prev_move = c[i - 1] - c[i - 5] if i >= 5 else 0
            if prev_move > 0:
                entry = c[i]; stop = entry + atr * 1.5; risk = stop - entry
                target = entry - risk * rr; d = -1
            else:
                entry = c[i]; stop = entry - atr * 1.5; risk = entry - stop
                target = entry + risk * rr; d = 1
            if risk <= 0: continue
            r = _walk_trade(df, i, entry, stop, target, d, max_bars)
            if r: results.append(r)
    return results


# ── Confluence / Multi-Factor Signal Matching ──

def _signal_index(trades: list[dict]) -> set:
    """Extract (entry_time, direction) for each trade as a set for fast lookup."""
    return {(t["entry_time"], t["direction"]) for t in trades}


def _confluence_filter(primary_trades: list[dict], confirmation_sets: list[set],
                       min_confirmations: int = 1) -> list[dict]:
    """Return primary trades where at least `min_confirmations` confirmation sets
    agree on direction at the same entry bar (±0 tolerance)."""
    if not confirmation_sets:
        return primary_trades
    confirmed = []
    for t in primary_trades:
        key = (t["entry_time"], t["direction"])
        matches = sum(1 for cs in confirmation_sets if key in cs)
        if matches >= min_confirmations:
            confirmed.append(t)
    return confirmed


# ── Multi-Factor Confluence Strategies ──

def backtest_vwap_momentum_confluence(df, rr=1.5, max_bars=20):
    """VWAP Reversion with Momentum Breakout confluence.
    Only take VWAP reversion trades where Momentum also signals same direction."""
    primary = backtest_vwap_reversion(df, rr=rr, max_bars=max_bars)
    conf_trades = backtest_momentum_breakout(df, rr=rr, max_bars=max_bars)
    return _confluence_filter(primary, [_signal_index(conf_trades)], min_confirmations=1)


def backtest_vwap_quote_confluence(df, rr=1.5, max_bars=20):
    """VWAP Reversion with Quote Imbalance Breakout confluence."""
    primary = backtest_vwap_reversion(df, rr=rr, max_bars=max_bars)
    conf_trades = backtest_quote_imbalance_breakout(df, rr=rr, max_bars=max_bars)
    return _confluence_filter(primary, [_signal_index(conf_trades)], min_confirmations=1)


def backtest_vwap_delta_confluence(df, rr=1.5, max_bars=20):
    """VWAP Reversion with Delta Confirmation Breakout confluence."""
    primary = backtest_vwap_reversion(df, rr=rr, max_bars=max_bars)
    conf_trades = backtest_delta_breakout(df, rr=rr, max_bars=max_bars)
    return _confluence_filter(primary, [_signal_index(conf_trades)], min_confirmations=1)


def backtest_vwap_triple_confluence(df, rr=1.5, max_bars=20):
    """VWAP Reversion requiring ≥2 of [Momentum, Quote Imbalance, Delta] to agree."""
    primary = backtest_vwap_reversion(df, rr=rr, max_bars=max_bars)
    conf_sets = [
        _signal_index(backtest_momentum_breakout(df, rr=rr, max_bars=max_bars)),
        _signal_index(backtest_quote_imbalance_breakout(df, rr=rr, max_bars=max_bars)),
        _signal_index(backtest_delta_breakout(df, rr=rr, max_bars=max_bars)),
    ]
    return _confluence_filter(primary, conf_sets, min_confirmations=2)


def backtest_rsi_momentum_confluence(df, rr=1.5, max_bars=15):
    """RSI Scalp with Momentum Breakout confluence."""
    primary = backtest_rsi_scalp(df, rr=rr, max_bars=max_bars)
    conf_trades = backtest_momentum_breakout(df, rr=rr, max_bars=max_bars)
    return _confluence_filter(primary, [_signal_index(conf_trades)], min_confirmations=1)


def backtest_rsi_delta_confluence(df, rr=1.5, max_bars=15):
    """RSI Scalp with Delta Confirmation Breakout confluence."""
    primary = backtest_rsi_scalp(df, rr=rr, max_bars=max_bars)
    conf_trades = backtest_delta_breakout(df, rr=rr, max_bars=max_bars)
    return _confluence_filter(primary, [_signal_index(conf_trades)], min_confirmations=1)


def backtest_ema_momentum_confluence(df, rr=1.5, max_bars=15):
    """EMA Pullback with Momentum Breakout confluence."""
    primary = backtest_ema_pullback(df, rr=rr, max_bars=max_bars)
    conf_trades = backtest_momentum_breakout(df, rr=rr, max_bars=max_bars)
    return _confluence_filter(primary, [_signal_index(conf_trades)], min_confirmations=1)


def backtest_ema_delta_confluence(df, rr=1.5, max_bars=15):
    """EMA Pullback with Delta Confirmation Breakout confluence."""
    primary = backtest_ema_pullback(df, rr=rr, max_bars=max_bars)
    conf_trades = backtest_delta_breakout(df, rr=rr, max_bars=max_bars)
    return _confluence_filter(primary, [_signal_index(conf_trades)], min_confirmations=1)


def backtest_orb_volume_confluence(df, rr=1.5, max_bars=30):
    """Opening Range Breakout with Liquidity Vacuum confluence."""
    primary = backtest_opening_range_breakout(df, rr=rr, max_bars=max_bars)
    conf_trades = backtest_liquidity_vacuum(df, rr=rr, max_bars=max_bars)
    return _confluence_filter(primary, [_signal_index(conf_trades)], min_confirmations=1)


def backtest_momentum_quote_confluence(df, rr=1.5, max_bars=20):
    """Momentum Breakout with Quote Imbalance Breakout confluence."""
    primary = backtest_momentum_breakout(df, rr=rr, max_bars=max_bars)
    conf_trades = backtest_quote_imbalance_breakout(df, rr=rr, max_bars=max_bars)
    return _confluence_filter(primary, [_signal_index(conf_trades)], min_confirmations=1)


def backtest_momentum_delta_quote_confluence(df, rr=1.5, max_bars=20):
    """Momentum Breakout requiring ≥2 of [Delta, Quote Imbalance, VWAP+Delta] to confirm."""
    primary = backtest_momentum_breakout(df, rr=rr, max_bars=max_bars)
    conf_sets = [
        _signal_index(backtest_delta_breakout(df, rr=rr, max_bars=max_bars)),
        _signal_index(backtest_quote_imbalance_breakout(df, rr=rr, max_bars=max_bars)),
        _signal_index(backtest_vwap_delta(df, rr=rr, max_bars=max_bars)),
    ]
    return _confluence_filter(primary, conf_sets, min_confirmations=2)


STRATEGIES = {
    "Momentum Breakout": backtest_momentum_breakout,
    "VWAP Reversion": backtest_vwap_reversion,
    "RSI Scalp": backtest_rsi_scalp,
    "EMA Pullback": backtest_ema_pullback,
    "Opening Range Breakout": backtest_opening_range_breakout,
    "Delta Confirmation Breakout": backtest_delta_breakout,
    "Quote Imbalance Breakout": backtest_quote_imbalance_breakout,
    "VWAP + Delta": backtest_vwap_delta,
    "Liquidity Vacuum": backtest_liquidity_vacuum,
    "VWAP + Momentum Confluence": backtest_vwap_momentum_confluence,
    "VWAP + Quote Imbalance Confluence": backtest_vwap_quote_confluence,
    "VWAP + Delta Confluence": backtest_vwap_delta_confluence,
    "VWAP Triple Confluence": backtest_vwap_triple_confluence,
    "RSI + Momentum Confluence": backtest_rsi_momentum_confluence,
    "RSI + Delta Confluence": backtest_rsi_delta_confluence,
    "EMA + Momentum Confluence": backtest_ema_momentum_confluence,
    "EMA + Delta Confluence": backtest_ema_delta_confluence,
    "ORB + Volume Confluence": backtest_orb_volume_confluence,
    "Momentum + Quote Confluence": backtest_momentum_quote_confluence,
    "Momentum + Delta + Quote Confluence": backtest_momentum_delta_quote_confluence,
}


def select_strategies(daily_row: pd.Series, enabled: list[str] | None = None) -> list[str]:
    """Adaptive strategy selection based on the stock's daily characteristics.
    If `enabled` is provided, only select from those strategies.
    If the user explicitly specified strategies via --strategies, use all of them
    regardless of score thresholds (explicit user intent bypasses adaptive filtering)."""
    available = STRATEGIES if enabled is None else [s for s in STRATEGIES if s in enabled]
    if not available:
        available = list(STRATEGIES)

    selected = []
    of_score = daily_row.get("orderflow_score", 50)
    qu_score = daily_row.get("quotes_score", 50)
    vol_score = daily_row.get("volatility_score", 50)
    liq_score = daily_row.get("liquidity_score", 50)

    def pick(name):
        if name in available and name not in selected:
            selected.append(name)

    # If user explicitly specified strategies, use ALL of them unconditionally
    if enabled:
        for s in enabled:
            pick(s)
        return selected

    pick("VWAP Reversion")
    pick("RSI Scalp")

    if vol_score > 55:
        pick("Momentum Breakout")
        pick("EMA Pullback")
        pick("Opening Range Breakout")

    if of_score > 50:
        pick("Delta Confirmation Breakout")
        pick("VWAP + Delta")

    if qu_score > 50:
        pick("Quote Imbalance Breakout")

    if of_score > 50 and liq_score > 50:
        pick("Liquidity Vacuum")

    # ── Confluence strategies: require both data sources to be strong ──

    # VWAP x Momentum: volatility + order flow both confirm
    if vol_score > 55 and of_score > 50:
        pick("VWAP + Momentum Confluence")

    # VWAP x Quote Imbalance: volatility + quotes both strong
    if vol_score > 55 and qu_score > 50:
        pick("VWAP + Quote Imbalance Confluence")

    # VWAP x Delta: volatility + order flow
    if vol_score > 55 and of_score > 50:
        pick("VWAP + Delta Confluence")

    # Triple confluence (VWAP): requires strong vol + order flow + quotes
    if vol_score > 55 and of_score > 50 and qu_score > 50:
        pick("VWAP Triple Confluence")

    # RSI x Momentum
    if vol_score > 55 and of_score > 50:
        pick("RSI + Momentum Confluence")

    # RSI x Delta
    if of_score > 50:
        pick("RSI + Delta Confluence")

    # EMA x Momentum
    if vol_score > 55 and of_score > 50:
        pick("EMA + Momentum Confluence")

    # EMA x Delta
    if vol_score > 55 and of_score > 50:
        pick("EMA + Delta Confluence")

    # ORB x Volume (Liquidity Vacuum)
    if vol_score > 55 and liq_score > 50:
        pick("ORB + Volume Confluence")

    # Momentum x Quote
    if vol_score > 55 and qu_score > 50:
        pick("Momentum + Quote Confluence")

    # Triple confirmation on Momentum
    if vol_score > 55 and of_score > 50 and qu_score > 50:
        pick("Momentum + Delta + Quote Confluence")

    return selected


# ══════════════════════════════════════════════════════════════════════
# SECTION 6b: Multiprocessing Worker Functions (module-level for pickle)
# ══════════════════════════════════════════════════════════════════════

def _features_worker(args: tuple) -> tuple:
    """Worker: compute daily features for one ticker. Returns (ticker, daily_df)."""
    ticker, year, agg, date_range = args
    _, daily_df = compute_daily_features(ticker, year, agg, date_range)
    return ticker, daily_df


def _backtest_worker(args: tuple) -> dict:
    """Worker: backtest one ticker across all strategies & R:R ratios.
    Must be module-level for pickle. Returns a dict of results."""
    (ticker, year, agg, date_range, avg_row_dict, enabled_strategies, rr_ratios,
     num_trades, capital_kwargs, time_zones_list) = args

    cap = CapitalManager(**capital_kwargs)
    time_zones = [(n, (sh, sm), (eh, em)) for n, (sh, sm), (eh, em) in time_zones_list]

    df, _ = compute_daily_features(ticker, year, agg, date_range)
    if df.empty:
        return {"ticker": ticker,
                "summaries": [], "portfolio": [], "trades": [],
                "time_zone_pnls": [], "strategies_used": []}

    avg_row = pd.Series(avg_row_dict)
    strategies = select_strategies(avg_row, enabled_strategies)

    summaries = []
    portfolio = []
    trades_out = []
    tz_pnls = []

    for s_name in strategies:
        if s_name not in STRATEGIES:
            continue
        s_func = STRATEGIES[s_name]
        for rr in rr_ratios:
            cap.reset()
            raw_trades = s_func(df, rr=rr)
            if num_trades and num_trades > 0 and len(raw_trades) > num_trades:
                raw_trades = raw_trades[:num_trades]
            if not raw_trades:
                continue

            summary = summarize_trades(raw_trades, s_name, ticker, rr)
            summaries.append(summary)

            port_trades = []
            for t in raw_trades:
                d = 1 if t["direction"] == "long" else -1
                res = cap.execute_trade(t["entry_price"], t["exit_price"],
                                         t.get("stop_price", t["entry_price"]),
                                         d, t["exit_time"])
                if res:
                    et = t["entry_time"]
                    if hasattr(et, "tzinfo") and et.tzinfo is not None:
                        et = et.replace(tzinfo=None)
                    ext = t["exit_time"]
                    if hasattr(ext, "tzinfo") and ext.tzinfo is not None:
                        ext = ext.replace(tzinfo=None)
                    port_trades.append({**t, **res, "ticker": ticker,
                                          "entry_time": et, "exit_time": ext,
                                          "strategy": s_name, "rr_ratio": rr})
                    tz_pnls.append((s_name, rr,
                                     time_zone_name(t["entry_time"], time_zones),
                                     t["pnl_%"]))
            if port_trades:
                trades_out.extend(port_trades)

            ps = {
                "ticker": ticker, "strategy": s_name, "rr_ratio": rr,
                "starting_cash": cap.starting_cash,
                "final_equity": round(cap.equity, 2),
                "total_return_$": round(cap.equity - cap.starting_cash, 2),
                "total_return_%": round(cap.total_return_pct(), 2),
                "max_drawdown_%": round(
                    max((p["drawdown_pct"] for p in cap.equity_curve), default=0), 2),
                "total_trades": cap.trade_count,
                "total_commission": round(cap.total_commission, 2),
            }
            portfolio.append(ps)

    return {"ticker": ticker,
            "summaries": summaries, "portfolio": portfolio, "trades": trades_out,
            "time_zone_pnls": tz_pnls, "strategies_used": strategies}


# ══════════════════════════════════════════════════════════════════════
# SECTION 6c: Reporting Helpers
# ══════════════════════════════════════════════════════════════════════

def _fmt_pct(v, decimals=1):
    if pd.isna(v) or not np.isfinite(v): return "N/A"
    return f"{v:.{decimals}f}%"


def _fmt_dollar(v, decimals=0):
    if pd.isna(v) or not np.isfinite(v): return "N/A"
    return f"${v:+,.{decimals}f}"


def _fmt_num(v, decimals=2):
    if pd.isna(v) or not np.isfinite(v): return "N/A"
    return f"{v:.{decimals}f}"


def _safe_div(a, b, default=0):
    return a / b if b else default


def _trade_metrics(trades_df):
    """Compute comprehensive trade statistics from a DataFrame of trades."""
    if trades_df.empty:
        return {}
    pnls = trades_df["pnl_$"].dropna()
    wins = trades_df[trades_df["result"] == "win"]
    losses = trades_df[trades_df["result"] == "loss"]
    n, nw, nl = len(trades_df), len(wins), len(losses)
    wr = _safe_div(nw, nw + nl) * 100
    total_win = wins["pnl_$"].sum() if nw > 0 else 0
    total_loss = abs(losses["pnl_$"].sum()) if nl > 0 else 0
    pf = _safe_div(total_win, total_loss, 999 if total_win > 0 else 0)
    total_pnl = pnls.sum()
    avg_pnl = pnls.mean()
    avg_win = wins["pnl_$"].mean() if nw > 0 else 0
    avg_loss = losses["pnl_$"].mean() if nl > 0 else 0
    largest_win = wins["pnl_$"].max() if nw > 0 else 0
    largest_loss = losses["pnl_$"].min() if nl > 0 else 0
    std_pnl = pnls.std()
    exp = pnls.mean()
    sharpe = _safe_div(exp, std_pnl) * np.sqrt(252) if std_pnl and std_pnl > 0 else 0
    ci95 = 1.96 * std_pnl / np.sqrt(n) if n > 1 and std_pnl > 0 else 0
    running = max_cons = 0
    for r in trades_df["result"]:
        if r in ("loss", "timeout"):
            running += 1; max_cons = max(max_cons, running)
        else:
            running = 0
    return {
        "n_trades": n, "n_wins": nw, "n_losses": nl,
        "win_rate": wr, "profit_factor": pf, "total_pnl": total_pnl,
        "avg_pnl": avg_pnl, "avg_win": avg_win, "avg_loss": avg_loss,
        "largest_win": largest_win, "largest_loss": largest_loss,
        "std_pnl": std_pnl, "sharpe": sharpe, "expectancy": exp,
        "ci95_expectancy": ci95, "max_consec_losses": max_cons,
        "win_loss_ratio": _safe_div(abs(avg_win), abs(avg_loss)) if avg_loss else 0,
        "median_pnl": pnls.median(),
        "profitable_days_pct": 0,
    }


def _excel_sheet(rows, columns, formats=None):
    return {"rows": rows, "columns": columns, "formats": formats or {}}


def _write_sheet(ws, data, wb=None):
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import numbers
    if isinstance(data, dict) and "rows" in data:
        rows_data = data["rows"]
        columns = data["columns"]
        formats = data.get("formats", {})
    elif isinstance(data, pd.DataFrame):
        return _write_sheet(ws, {
            "rows": [row for row in dataframe_to_rows(data, index=False, header=True)],
            "columns": list(data.columns),
        })
    else:
        return

    for ri, row in enumerate(rows_data, 1):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if formats:
                col_name = columns[ci - 1] if ci <= len(columns) else None
                fmt = formats.get(col_name)
                if fmt == "pct" and isinstance(val, (int, float)) and ri > 1:
                    cell.number_format = '0.00%'
                    cell.value = val / 100.0 if abs(val) > 1 else val
                elif fmt == "dollar" and isinstance(val, (int, float)):
                    cell.number_format = '$#,##0.00'
                elif fmt == "int":
                    cell.number_format = '#,##0'


def _apply_formatting(ws, columns, hdr_fill, hdr_font, border, alignment, num_cols=None):
    from openpyxl.styles import Font, PatternFill, Alignment, Border
    mc = num_cols or len(columns) or ws.max_column
    for c in range(1, mc + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = alignment
        cell.border = border
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _add_conditional_format(ws, col_idx, col_name):
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import PatternFill
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    col_letter = _openpyxl_col_letter(col_idx)
    rng = f"{col_letter}2:{col_letter}{ws.max_row}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["0"], fill=green))
    ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0"], fill=red))


def _openpyxl_col_letter(idx):
    from openpyxl.utils import get_column_letter
    return get_column_letter(idx)


# ══════════════════════════════════════════════════════════════════════
# SECTION 7: Main Pipeline
# ══════════════════════════════════════════════════════════════════════

def summarize_trades(trades: list[dict], strategy: str, ticker: str, rr: float) -> dict:
    if not trades:
        return {"strategy": strategy, "ticker": ticker, "rr_ratio": rr, "total_trades": 0,
                "wins": 0, "losses": 0, "timeouts": 0, "win_rate_%": 0, "profit_factor": 0,
                "total_pnl_%": 0, "avg_win_%": 0, "avg_loss_%": 0, "expectancy_%": 0,
                "sharpe_approx": 0, "avg_bars_held": 0, "max_consec_losses": 0}
    df = pd.DataFrame(trades)
    wins = df[df["result"] == "win"]
    losses = df[df["result"] == "loss"]
    n, nw, nl = len(df), len(wins), len(losses)
    wr = nw / (nw + nl) * 100 if (nw + nl) > 0 else 0
    total_win = wins["pnl_%"].sum() if nw > 0 else 0
    total_loss = abs(losses["pnl_%"].sum()) if nl > 0 else 0
    pf = total_win / total_loss if total_loss > 0 else (999 if total_win > 0 else 0)
    total_pnl = df["pnl_%"].sum()
    exp = total_pnl / n if n > 0 else 0
    std = df["pnl_%"].std()
    sharpe = (exp / std * np.sqrt(252 / 20)) if std and std > 0 else 0
    running = max_cons = 0
    for r in df["result"]:
        if r in ("loss", "timeout"):
            running += 1; max_cons = max(max_cons, running)
        else:
            running = 0
    return {
        "strategy": strategy, "ticker": ticker, "rr_ratio": rr,
        "total_trades": n, "wins": nw, "losses": nl, "timeouts": n - nw - nl,
        "win_rate_%": round(wr, 1), "profit_factor": round(pf, 2),
        "total_pnl_%": round(total_pnl, 4), "avg_win_%": round(wins["pnl_%"].mean(), 4) if nw > 0 else 0,
        "avg_loss_%": round(abs(losses["pnl_%"].mean()), 4) if nl > 0 else 0,
        "expectancy_%": round(exp, 4), "sharpe_approx": round(sharpe, 2),
        "avg_bars_held": round(df["bars_held"].mean(), 1) if n > 0 else 0,
        "max_consec_losses": max_cons,
    }


def time_zone_name(ts, tz_map):
    try:
        et = ts.tz_convert("US/Eastern")
        h, m = et.hour, et.minute
    except Exception:
        return "Other"
    for name, (sh, sm), (eh, em) in tz_map:
        if (h > sh or (h == sh and m >= sm)) and (h < eh or (h == eh and m < em)):
            return name
    return "Other"


def _strip_tz(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            try:
                df[col] = df[col].dt.tz_localize(None)
            except Exception:
                pass
    return df


def run_pipeline(year: str, top_n: int, backtest_n: int, num_trades: int,
                 capital: CapitalManager, output_path: str,
                 enabled_strategies: list[str] | None = None,
                 rr_ratios: list[float] | None = None,
                 nprocs: int = 1,
                 aggregate: str = DEFAULT_AGGREGATE,
                 date_range: tuple | None = None):
    if rr_ratios is None:
        rr_ratios = RR_RATIOS
    actual_nprocs = max(1, min(nprocs, os.cpu_count() or 1))
    t_start = time.time()

    # ── Data Availability Report ──
    availability, counts = report_data_availability(year, aggregate)
    log.info("Data Availability Report:")
    for k, v in counts.items():
        log.info("  %s: %d stocks", k, v)
    all_tickers = sorted(availability.keys())
    log.info("  Total unique tickers with OHLCV: %d", len(all_tickers))

    # Report missing data
    missing = {}
    for col in ["Trades", "Quotes", "Options", "Chains"]:
        missing[col] = [t for t, d in availability.items() if not d.get(col)]
        if missing[col]:
            log.warning("  Missing %s for %d tickers: %s...",
                        col, len(missing[col]), ", ".join(missing[col][:5]))

    # ── Daily Feature Computation & Ranking ──
    target_tickers = all_tickers[:top_n]
    log.info("Computing daily features for %d tickers...", len(target_tickers))
    all_daily = []

    if actual_nprocs > 1:
        work_items = [(t, year, aggregate, date_range) for t in target_tickers]
        log.info("  Using %d processes for feature computation", actual_nprocs)
        with ProcessPoolExecutor(max_workers=actual_nprocs) as pool:
            futures = {pool.submit(_features_worker, w): w[0] for w in work_items}
            completed = 0
            for fut in as_completed(futures):
                ticker = futures[fut]
                try:
                    _, daily_df = fut.result()
                    if not daily_df.empty:
                        all_daily.append(daily_df)
                except Exception as e:
                    log.error("  Error computing features for %s: %s", ticker, e)
                completed += 1
                if completed % 10 == 0:
                    log.info("  Processed %d/%d tickers", completed, len(target_tickers))
    else:
        for i, ticker in enumerate(target_tickers):
            _, daily_df = compute_daily_features(ticker, year, aggregate, date_range)
            if not daily_df.empty:
                all_daily.append(daily_df)
            if (i + 1) % 10 == 0:
                log.info("  Processed %d/%d tickers", i + 1, len(target_tickers))

    if not all_daily:
        log.error("No daily data computed. Exiting.")
        return

    combined_daily = pd.concat(all_daily, ignore_index=True)
    combined_daily = compute_opportunity_score(combined_daily)

    # ── Daily ranking ──
    log.info("Ranking stocks per day...")
    daily_rankings = []
    for date, grp in combined_daily.groupby("date"):
        ranked = grp.sort_values("opportunity_score", ascending=False)
        ranked["daily_rank"] = range(1, len(ranked) + 1)
        daily_rankings.append(ranked)
    daily_df = pd.concat(daily_rankings, ignore_index=True) if daily_rankings else combined_daily

    backtest_tickers = daily_df.sort_values("opportunity_score", ascending=False)["ticker"].unique()[:backtest_n]

    # ── Backtest ──
    log.info("Backtesting %d tickers with adaptive strategy selection...", len(backtest_tickers))
    all_trades = []
    all_summaries = []
    all_portfolio = []
    time_zone_stats: dict[str, list] = defaultdict(list)
    strategy_usage: dict[str, int] = defaultdict(int)

    # Build serializable capital params and time zones
    capital_kwargs = {
        "starting_cash": capital.starting_cash,
        "risk_pct": capital.risk_pct,
        "compounding": capital.compounding,
        "margin_rate": capital.margin_rate,
        "max_position_pct": capital.max_position_pct,
    }
    tz_serializable = [(n, (sh, sm), (eh, em)) for n, (sh, sm), (eh, em) in TIME_ZONES]

    if actual_nprocs > 1:
        log.info("  Using %d processes for backtesting", actual_nprocs)
        work_items = []
        for ticker in backtest_tickers:
            ticker_daily = daily_df[daily_df["ticker"] == ticker]
            avg_row_dict = (ticker_daily.mean(numeric_only=True).to_dict()
                            if not ticker_daily.empty else {})
            work_items.append((
                ticker, year, aggregate, date_range, avg_row_dict, enabled_strategies, rr_ratios,
                num_trades, capital_kwargs, tz_serializable,
            ))
        with ProcessPoolExecutor(max_workers=actual_nprocs) as pool:
            futures = {pool.submit(_backtest_worker, w): w[0] for w in work_items}
            completed = 0
            for fut in as_completed(futures):
                ticker = futures[fut]
                try:
                    result = fut.result()
                    all_summaries.extend(result["summaries"])
                    all_portfolio.extend(result["portfolio"])
                    all_trades.extend(result["trades"])
                    for s_name, rr, tz, pnl in result["time_zone_pnls"]:
                        time_zone_stats[(s_name, rr, tz)].append(pnl)
                    for s in result["strategies_used"]:
                        strategy_usage[s] += 1
                    log.info("  Backtested %s with strategies: %s", ticker,
                             ", ".join(result["strategies_used"][:4]))
                except Exception as e:
                    log.error("  Error backtesting %s: %s", ticker, e)
                completed += 1
    else:
        for ticker in backtest_tickers:
            df, _ = compute_daily_features(ticker, year, aggregate, date_range)
            if df.empty:
                continue

            ticker_daily = daily_df[daily_df["ticker"] == ticker]
            avg_row = ticker_daily.mean(numeric_only=True) if not ticker_daily.empty else pd.Series()
            strategies = select_strategies(avg_row, enabled_strategies)

            for s_name in strategies:
                strategy_usage[s_name] += 1
                if s_name not in STRATEGIES:
                    continue
                s_func = STRATEGIES[s_name]
                for rr in rr_ratios:
                    capital.reset()
                    trades = s_func(df, rr=rr)
                    if num_trades and num_trades > 0 and len(trades) > num_trades:
                        trades = trades[:num_trades]
                    if not trades:
                        continue

                    summary = summarize_trades(trades, s_name, ticker, rr)
                    all_summaries.append(summary)

                    port_trades = []
                    for t in trades:
                        d = 1 if t["direction"] == "long" else -1
                        res = capital.execute_trade(t["entry_price"], t["exit_price"],
                                                     t.get("stop_price", t["entry_price"]),
                                                     d, t["exit_time"])
                        if res:
                            et = t["entry_time"]
                            if hasattr(et, "tzinfo") and et.tzinfo is not None:
                                et = et.replace(tzinfo=None)
                            ext = t["exit_time"]
                            if hasattr(ext, "tzinfo") and ext.tzinfo is not None:
                                ext = ext.replace(tzinfo=None)
                            port_trades.append({**t, **res, "ticker": ticker,
                                                  "entry_time": et, "exit_time": ext,
                                                  "strategy": s_name, "rr_ratio": rr})
                            tz_name = time_zone_name(t["entry_time"], TIME_ZONES)
                            time_zone_stats[(s_name, rr, tz_name)].append(t["pnl_%"])
                    if port_trades:
                        all_trades.extend(port_trades)

                    ps = {
                        "ticker": ticker, "strategy": s_name, "rr_ratio": rr,
                        "starting_cash": capital.starting_cash,
                        "final_equity": round(capital.equity, 2),
                        "total_return_$": round(capital.equity - capital.starting_cash, 2),
                        "total_return_%": round(capital.total_return_pct(), 2),
                        "max_drawdown_%": round(max((p["drawdown_pct"] for p in capital.equity_curve), default=0), 2),
                        "total_trades": capital.trade_count,
                        "total_commission": round(capital.total_commission, 2),
                    }
                    all_portfolio.append(ps)

            log.info("  Backtested %s with strategies: %s", ticker, ", ".join(strategies[:4]))

    # ── Strategy usage report ──
    log.info("Strategy usage across tickers:")
    for s, count in sorted(strategy_usage.items(), key=lambda x: -x[1]):
        log.info("  %s: %d selections", s, count)

    #     ── Build analysis DataFrames ──
    sum_df = pd.DataFrame(all_summaries)
    port_df = pd.DataFrame(all_portfolio)
    trade_df = pd.DataFrame(all_trades) if all_trades else pd.DataFrame()
    lb = []; slb = []  # default empty
    if not trade_df.empty:
        trade_df["entry_time_dt"] = pd.to_datetime(trade_df["entry_time"])
        trade_df["exit_time_dt"] = pd.to_datetime(trade_df["exit_time"])
        trade_df["day_of_week"] = trade_df["entry_time_dt"].dt.day_name()
        trade_df["pnl_$"] = pd.to_numeric(trade_df.get("pnl_$", 0), errors="coerce").fillna(0)

    elapsed = time.time() - t_start
    mins, secs = divmod(int(elapsed), 60)

    # ═══════════════════════════════════════════════════════════════════
    # STDOUT: Comprehensive Summary Tables
    # ═══════════════════════════════════════════════════════════════════

    W = 90  # table width

    # ── A. Overall Strategy × R:R Leaderboard ──
    print("\n" + "=" * W)
    print("  A. OVERALL STRATEGY × R:R LEADERBOARD")
    print("=" * W)
    if not sum_df.empty:
        leaderboard_data = []
        for (s_name, rr), grp in sum_df.groupby(["strategy", "rr_ratio"]):
            trades_sub = trade_df[(trade_df["strategy"] == s_name) & (trade_df["rr_ratio"] == rr)] if not trade_df.empty else pd.DataFrame()
            m = _trade_metrics(trades_sub)
            agg_pnl_pct = grp["total_pnl_%"].sum()
            leaderboard_data.append({
                "Strategy": s_name, "R:R": f"1:{rr}",
                "Trades": m.get("n_trades", grp["total_trades"].sum()),
                "WinRate": m.get("win_rate", 0), "PF": m.get("profit_factor", 0),
                "Expectancy$": m.get("expectancy", 0), "TotalPnL$": m.get("total_pnl", 0),
                "TotalPnL%": agg_pnl_pct, "Sharpe": m.get("sharpe", 0),
                "MaxDD%": port_df[(port_df["strategy"] == s_name) & (port_df["rr_ratio"] == rr)]["max_drawdown_%"].mean() if not port_df.empty else 0,
                "AvgWin$": m.get("avg_win", 0), "AvgLoss$": m.get("avg_loss", 0),
                "W/LRatio": m.get("win_loss_ratio", 0),
                "LargestWin$": m.get("largest_win", 0), "LargestLoss$": m.get("largest_loss", 0),
                "Samples": f"{m.get('n_trades', 0)} ({'OK' if m.get('n_trades', 0) >= 20 else 'LOW'})",
                "_score": (m.get("profit_factor", 0) * 3 +
                           m.get("sharpe", 0) * 2 +
                           (m.get("win_rate", 0) / 100) * 10),
            })
        lb = sorted(leaderboard_data, key=lambda x: x["_score"], reverse=True)
        hdr = f"{'Strategy':<22s} {'R:R':<8s} {'Trades':>7s} {'WR%':>7s} {'PF':>7s} {'Exp$':>8s} {'Total$':>10s} {'P&L%':>8s} {'Sharpe':>7s} {'MaxDD':>7s}"
        print(f"  {hdr}")
        print(f"  {'-' * len(hdr)}")
        for row in lb[:20]:
            print(f"  {row['Strategy']:<22s} {row['R:R']:<8s} {row['Trades']:>7d} {row['WinRate']:>6.1f}% {row['PF']:>6.2f} {row['Expectancy$']:>+7.2f} {row['TotalPnL$']:>+9.0f} {row['TotalPnL%']:>+7.2f}% {row['Sharpe']:>6.2f} {row['MaxDD%']:>6.1f}%")
        print()

    # ── B. Best Strategy ──
    print("-" * W)
    print("  B. BEST STRATEGY")
    print("-" * W)
    if not sum_df.empty:
        best = lb[0] if lb else {}
        print(f"  Best: {best.get('Strategy','N/A')} @ R:R {best.get('R:R','N/A')}")
        print(f"  Rationale: PF={best.get('PF',0):.2f}, Sharpe={best.get('Sharpe',0):.2f}, "
              f"WinRate={best.get('WinRate',0):.1f}% across {best.get('Trades',0)} trades")
        print(f"  Total P&L: {_fmt_dollar(best.get('TotalPnL$',0))}, "
              f"Max DD: {best.get('MaxDD%',0):.1f}%, Avg Win: {_fmt_dollar(best.get('AvgWin$',0))}, "
              f"Avg Loss: {_fmt_dollar(best.get('AvgLoss$',0))}")
        print(f"  Sample: {best.get('Samples','N/A')}")
    print()

    # ── C. Strategy Comparison ──
    print("-" * W)
    print("  C. STRATEGY COMPARISON")
    print("-" * W)
    if not sum_df.empty:
        hdr_c = f"{'Strategy':<24s} {'BestRR':<8s} {'WR%':>7s} {'PF':>7s} {'Exp$':>8s} {'P&L$':>10s} {'Sharpe':>7s} {'MaxDD':>7s} {'Trades':>7s}"
        print(f"  {hdr_c}")
        print(f"  {'-' * len(hdr_c)}")
        for s_name in sorted(sum_df["strategy"].unique()):
            s_lb = [r for r in lb if r["Strategy"] == s_name]
            if not s_lb: continue
            top = s_lb[0]
            print(f"  {s_name:<24s} {top['R:R']:<8s} {top['WinRate']:>6.1f}% {top['PF']:>6.2f} {top['Expectancy$']:>+7.2f} {top['TotalPnL$']:>+9.0f} {top['Sharpe']:>6.2f} {top['MaxDD%']:>6.1f}% {top['Trades']:>7d}")
    print()

    # ── D. R:R Comparison ──
    print("-" * W)
    print("  D. R:R COMPARISON")
    print("-" * W)
    if not sum_df.empty:
        hdr_d = f"{'R:R':<8s} {'Trades':>8s} {'WR%':>7s} {'PF':>7s} {'Exp$':>8s} {'P&L$':>10s} {'Sharpe':>7s} {'MaxDD':>7s}"
        print(f"  {hdr_d}")
        print(f"  {'-' * len(hdr_d)}")
        for rr in rr_ratios:
            rr_lb = [r for r in lb if r["R:R"] == f"1:{rr}"]
            if not rr_lb: continue
            top_rr = max(rr_lb, key=lambda x: x["_score"])
            ttl_trades = sum(r["Trades"] for r in rr_lb)
            ttl_pnl = sum(r["TotalPnL$"] for r in rr_lb)
            avg_wr = np.mean([r["WinRate"] for r in rr_lb]) if rr_lb else 0
            avg_pf = np.mean([r["PF"] for r in rr_lb]) if rr_lb else 0
            avg_exp = np.mean([r["Expectancy$"] for r in rr_lb]) if rr_lb else 0
            avg_sharpe = np.mean([r["Sharpe"] for r in rr_lb]) if rr_lb else 0
            avg_dd = np.mean([r["MaxDD%"] for r in rr_lb]) if rr_lb else 0
            print(f"  {f'1:{rr}':<8s} {ttl_trades:>8d} {avg_wr:>6.1f}% {avg_pf:>6.2f} {avg_exp:>+7.2f} {ttl_pnl:>+9.0f} {avg_sharpe:>6.2f} {avg_dd:>6.1f}%")
    print()

    # ── E. Stock Leaderboard ──
    print("-" * W)
    print("  E. STOCK LEADERBOARD")
    print("-" * W)
    if not sum_df.empty and not daily_df.empty:
        stock_metrics = []
        for ticker in sorted(sum_df["ticker"].unique()):
            tk_daily = daily_df[daily_df["ticker"] == ticker]
            avg_scores = tk_daily[["opportunity_score", "liquidity_score", "orderflow_score",
                                    "volatility_score", "quotes_score", "options_score",
                                    "cross_market_score"]].mean().to_dict() if not tk_daily.empty else {}
            tk_trades = trade_df[trade_df["ticker"] == ticker] if not trade_df.empty else pd.DataFrame()
            m = _trade_metrics(tk_trades)
            tk_sum = sum_df[sum_df["ticker"] == ticker]
            best_tk = tk_sum.sort_values("profit_factor", ascending=False).iloc[0] if not tk_sum.empty else {}

            composite = (m.get("profit_factor", 0) * 3 + m.get("sharpe", 0) * 2 -
                         m.get("max_consec_losses", 0) * 0.1 + m.get("win_rate", 0) / 50)

            stock_metrics.append({
                "Ticker": ticker,
                "OppScore": avg_scores.get("opportunity_score", 0),
                "Liq": avg_scores.get("liquidity_score", 0),
                "OF": avg_scores.get("orderflow_score", 0),
                "Vol": avg_scores.get("volatility_score", 0),
                "Qu": avg_scores.get("quotes_score", 0),
                "Opt": avg_scores.get("options_score", 0),
                "XMkt": avg_scores.get("cross_market_score", 0),
                "Strategy": best_tk.get("strategy", ""),
                "RR": best_tk.get("rr_ratio", 0),
                "Trades": m.get("n_trades", 0),
                "WR": m.get("win_rate", 0),
                "PF": m.get("profit_factor", 0),
                "Exp$": m.get("expectancy", 0),
                "PnL$": m.get("total_pnl", 0),
                "Sharpe": m.get("sharpe", 0),
                "MaxDD": best_tk.get("max_consec_losses", 0),
                "_composite": composite,
            })
        slb = sorted(stock_metrics, key=lambda x: x["_composite"], reverse=True)
        hdr_e = f"{'Ticker':<8s} {'Opp':>5s} {'Liq':>5s} {'OF':>5s} {'Vol':>5s} {'Qu':>5s} {'Opt':>5s} {'XMkt':>5s} {'Strat':<22s} {'RR':>6s} {'Trd':>5s} {'WR':>6s} {'PF':>6s} {'Exp$':>8s} {'PnL$':>10s} {'Shp':>6s}"
        print(f"  {hdr_e}")
        print(f"  {'-' * len(hdr_e)}")
        for row in slb[:30]:
            rr_str = f"1:{row['RR']}" if row['RR'] else "N/A"
            print(f"  {row['Ticker']:<8s} {row['OppScore']:>4.0f} {row['Liq']:>4.0f} {row['OF']:>4.0f} {row['Vol']:>4.0f} {row['Qu']:>4.0f} {row['Opt']:>4.0f} {row['XMkt']:>4.0f} {str(row['Strategy'])[:22]:<22s} {rr_str:>6s} {row['Trades']:>5d} {row['WR']:>5.1f}% {row['PF']:>5.2f} {row['Exp$']:>+7.2f} {row['PnL$']:>+9.0f} {row['Sharpe']:>5.2f}")
    print()

    # ── F. Best Trading Conditions ──
    print("-" * W)
    print("  F. BEST TRADING CONDITIONS")
    print("-" * W)
    if not trade_df.empty:
        # Time of day
        tod_stats = {}
        for tz_name in [t[0] for t in TIME_ZONES]:
            tz_mask = pd.Series(False, index=trade_df.index)
            for (s_name, rr, tz), _ in time_zone_stats.items():
                if tz == tz_name:
                    # cannot map back exactly, use approximate
                    pass
        print("  Time of Day (by pnl_% bucket):")
        for (s_name, rr, tz), pnls in sorted(time_zone_stats.items(), key=lambda x: -len(x[1]))[:8]:
            if len(pnls) < 5: continue
            avg_p = np.mean(pnls); wr = sum(1 for p in pnls if p > 0) / len(pnls) * 100
            print(f"    {tz:<14s} | {s_name:<22s} @ 1:{rr} | N={len(pnls):>4d}  AvgPnL={avg_p:+.4f}%  WR={wr:.1f}%")

        # Day of week
        if "day_of_week" in trade_df.columns:
            print("\n  Day of Week:")
            for dow in ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]:
                dow_t = trade_df[trade_df["day_of_week"] == dow]
                if len(dow_t) < 3: continue
                m = _trade_metrics(dow_t)
                print(f"    {dow:<11s} | N={m['n_trades']:>4d}  WR={m['win_rate']:.1f}%  PF={m['profit_factor']:.2f}  "
                      f"Exp={_fmt_dollar(m['expectancy'])}  P&L={_fmt_dollar(m['total_pnl'])}")

        # Volatility regime
        if not daily_df.empty:
            print("\n  Volatility Regime (ATR pct quintiles):")
            try:
                daily_df["_atr_q"] = pd.qcut(daily_df["atr_14_pct"].fillna(0), 5, labels=["V1","V2","V3","V4","V5"], duplicates="drop")
                for q in sorted(daily_df["_atr_q"].dropna().unique()):
                    q_dates = daily_df[daily_df["_atr_q"] == q][["ticker", "date"]]
                    if trade_df.empty:
                        break
                    q_trades = trade_df.merge(q_dates, on="ticker", how="inner") if not trade_df.empty else pd.DataFrame()
                    if len(q_trades) < 3: continue
                    m = _trade_metrics(q_trades)
                    print(f"    {q:<5s} | N={m['n_trades']:>4d}  WR={m['win_rate']:.1f}%  PF={m['profit_factor']:.2f}  "
                          f"Exp={_fmt_dollar(m['expectancy'])}  P&L={_fmt_dollar(m['total_pnl'])}")
            except Exception:
                pass

            # Liquidity regime
            print("\n  Liquidity Regime (liquidity score quintiles):")
            try:
                daily_df["_liq_q"] = pd.qcut(daily_df["liquidity_score"].fillna(50), 5, labels=["L1","L2","L3","L4","L5"], duplicates="drop")
                for q in sorted(daily_df["_liq_q"].dropna().unique()):
                    q_dates = daily_df[daily_df["_liq_q"] == q][["ticker", "date"]]
                    q_trades = trade_df.merge(q_dates, on="ticker", how="inner") if not trade_df.empty else pd.DataFrame()
                    if len(q_trades) < 3: continue
                    m = _trade_metrics(q_trades)
                    print(f"    {q:<5s} | N={m['n_trades']:>4d}  WR={m['win_rate']:.1f}%  PF={m['profit_factor']:.2f}  "
                          f"Exp={_fmt_dollar(m['expectancy'])}  P&L={_fmt_dollar(m['total_pnl'])}")
            except Exception:
                pass

    # ═══════════════════════════════════════════════════════════════════
    # EXCEL WORKBOOK
    # ═══════════════════════════════════════════════════════════════════

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.chart import LineChart, BarChart, Reference
    from openpyxl.chart.series import DataPoint
    import openpyxl as _openpyxl_mod

    wb = Workbook()
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    win_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    loss_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
    align_ctr = Alignment(horizontal="center", wrap_text=True)
    pct_fmt = '0.00%'
    dollar_fmt = '$#,##0.00'

    def _xl_hdr(ws, row=1, nc=None):
        nc = nc or ws.max_column
        for c in range(1, nc + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = align_ctr; cell.border = border
        ws.freeze_panes = "A2"
        try: ws.auto_filter.ref = ws.dimensions
        except Exception: pass

    def _xl_autow(ws, mw=40):
        for col in ws.columns:
            ml = max((min(len(str(c.value or "")), mw) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = ml + 2

    def _xl_df(ws, df, sort_by=None):
        if df.empty: return
        if sort_by:
            cols = [c for c in (sort_by if isinstance(sort_by, list) else [sort_by]) if c in df.columns]
            if cols:
                df = df.sort_values(cols, ascending=False)
        df = _strip_tz(df)
        for ri, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for ci, v in enumerate(row, 1): ws.cell(row=ri, column=ci, value=v)

    def _xl_cfmt(ws, col_idx):
        col_l = get_column_letter(col_idx)
        rng = f"{col_l}2:{col_l}{ws.max_row}"
        try:
            ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["0"],
                                         fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")))
            ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0"],
                                         fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")))
        except Exception: pass

    # ── Excel: Executive_Summary (Sheet 1) ──
    ws = wb.active; ws.title = "Executive_Summary"

    # single unified table: col A = metric labels, cols B+ = one per strategy (rank order)
    metric_labels = [
        "Total Trades", "Win Rate", "Profit Factor",
        "Expectancy/Trade", "Total P&L", "Total P&L %", "Sharpe Ratio",
        "Max Drawdown", "Avg Win", "Avg Loss", "Largest Win", "Largest Loss",
        "Sample Size", "Elapsed Time",
    ]

    def _fmt_entry(b, label):
        if label == "Total Trades": return b["Trades"]
        if label == "Win Rate": return f"{b['WinRate']:.1f}%"
        if label == "Profit Factor": return f"{b['PF']:.2f}"
        if label == "Expectancy/Trade": return f"${b['Expectancy$']:+.2f}"
        if label == "Total P&L": return f"${b['TotalPnL$']:+.0f}"
        if label == "Total P&L %": return f"{b['TotalPnL%']:+.2f}%"
        if label == "Sharpe Ratio": return f"{b['Sharpe']:.2f}"
        if label == "Max Drawdown": return f"{b['MaxDD%']:.1f}%"
        if label == "Avg Win": return _fmt_dollar(b['AvgWin$'])
        if label == "Avg Loss": return _fmt_dollar(b['AvgLoss$'])
        if label == "Largest Win": return _fmt_dollar(b['LargestWin$'])
        if label == "Largest Loss": return _fmt_dollar(b['LargestLoss$'])
        if label == "Sample Size": return b["Samples"]
        if label == "Elapsed Time": return f"{mins}m {secs}s"
        return ""

    # Row 1: column headers
    a1 = ws.cell(row=1, column=1, value="Metric")
    a1.font = hdr_font; a1.fill = hdr_fill; a1.alignment = align_ctr; a1.border = border
    for idx, b in enumerate(lb):
        col = idx + 2  # B=2, C=3, D=4, ...
        header_text = f"Rank {idx+1}\n{b['Strategy']}\n({b['R:R']})"
        cell = ws.cell(row=1, column=col, value=header_text)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = align_ctr; cell.border = border

    # Rows 2+: metric labels in col A, values in cols B+
    for ri, label in enumerate(metric_labels, 2):
        ws.cell(row=ri, column=1, value=label).border = border
        for idx, b in enumerate(lb):
            col = idx + 2
            cell = ws.cell(row=ri, column=col, value=_fmt_entry(b, label))
            cell.border = border
            cell.alignment = align_ctr

    metrics_end_row = 1 + len(metric_labels) + 1
    coverage_start = metrics_end_row + 1

    ws.cell(row=coverage_start, column=1, value="Data Coverage").font = hdr_font
    ws.cell(row=coverage_start, column=1).fill = hdr_fill
    ws.cell(row=coverage_start, column=1).border = border
    for i, (k, v) in enumerate(counts.items()):
        ri = coverage_start + 1 + i
        ws.cell(row=ri, column=1, value=f"  {k}").border = border
        ws.cell(row=ri, column=2, value=f"{v} stocks").border = border

    pos_start = coverage_start + 1 + len(counts) + 1
    ws.cell(row=pos_start, column=1, value="Position Sizing").font = hdr_font
    ws.cell(row=pos_start, column=1).fill = hdr_fill
    ws.cell(row=pos_start, column=1).border = border
    pos_items = [
        ("Risk per trade", f"{capital.risk_pct*100:.1f}% = ${capital._risk_dollars():,.0f}"),
        ("Max position %", f"{capital.max_position_pct*100:.0f}% of equity"),
        ("Margin rate", f"{capital.margin_rate:.0%}"),
    ]
    for i, (label, val) in enumerate(pos_items):
        ri = pos_start + 1 + i
        ws.cell(row=ri, column=1, value=label).border = border
        ws.cell(row=ri, column=2, value=val).border = border

    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 22
    for idx in range(len(lb)):
        col = idx + 2
        ws.column_dimensions[get_column_letter(col)].width = 30

    # ── Excel: Strategy_Comparison ──
    ws = wb.create_sheet("(C) Strategy_Comparison")
    sc_cols = ["Strategy", "R:R", "Trades", "WinRate", "PF", "Expectancy$", "TotalPnL$", "TotalPnL%", "Sharpe", "MaxDD%", "AvgWin$", "AvgLoss$", "LargestWin$", "LargestLoss$", "Sample"]
    sc_df = pd.DataFrame(lb)[[c for c in sc_cols if c in (lb[0] if lb else {})]] if lb else pd.DataFrame()
    _xl_df(ws, sc_df, sort_by=["PF", "WinRate"]); _xl_hdr(ws); _xl_autow(ws)

    # Charts: 4 separate charts — PF, WinRate, Sharpe+PnL%, Trades
    if not sc_df.empty and len(sc_df) > 0:
        n_rows = len(sc_df) + 1
        cats = Reference(ws, min_col=1, min_row=2, max_row=n_rows)
        chart_w = 44; chart_h = 20

        # Chart 1: Profit Factor
        ch1 = BarChart()
        ch1.title = "Profit Factor"; ch1.y_axis.title = "PF"
        ch1.style = 10; ch1.height = chart_h; ch1.width = chart_w
        ch1.add_data(Reference(ws, min_col=5, min_row=1, max_row=n_rows, max_col=5), titles_from_data=True)
        ch1.set_categories(cats)
        ch1.series[0].graphicalProperties.solidFill = "2F5496"
        ch1.legend = None
        ws.add_chart(ch1, f"A{n_rows + 3}")

        # Chart 2: Win Rate
        ch2 = BarChart()
        ch2.title = "Win Rate %"; ch2.y_axis.title = "Win Rate (%)"
        ch2.style = 10; ch2.height = chart_h; ch2.width = chart_w
        ch2.add_data(Reference(ws, min_col=4, min_row=1, max_row=n_rows, max_col=4), titles_from_data=True)
        ch2.set_categories(cats)
        ch2.series[0].graphicalProperties.solidFill = "4472C4"
        ch2.legend = None
        ws.add_chart(ch2, f"A{n_rows + chart_h + 5}")

        # Chart 3: Sharpe + Total PnL% (blue=Sharpe, gold=PnL%)
        ch3 = BarChart()
        ch3.grouping = "clustered"
        ch3.title = "Sharpe Ratio & Total PnL%"; ch3.y_axis.title = "Sharpe / PnL%"
        ch3.style = 10; ch3.height = chart_h; ch3.width = chart_w
        ch3.add_data(Reference(ws, min_col=9, min_row=1, max_row=n_rows, max_col=9), titles_from_data=True)
        ch3.add_data(Reference(ws, min_col=8, min_row=1, max_row=n_rows, max_col=8), titles_from_data=True)
        ch3.set_categories(cats)
        ch3.series[0].graphicalProperties.solidFill = "2F5496"
        ch3.series[1].graphicalProperties.solidFill = "FFC000"
        ws.add_chart(ch3, f"A{n_rows + chart_h * 2 + 7}")

        # Chart 4: Total Trades (sample size context)
        ch4 = BarChart()
        ch4.title = "Total Trades"; ch4.y_axis.title = "Trades"
        ch4.style = 10; ch4.height = chart_h; ch4.width = chart_w
        ch4.add_data(Reference(ws, min_col=3, min_row=1, max_row=n_rows, max_col=3), titles_from_data=True)
        ch4.set_categories(cats)
        ch4.series[0].graphicalProperties.solidFill = "7030A0"
        ch4.legend = None
        ws.add_chart(ch4, f"A{n_rows + chart_h * 3 + 9}")

    # ── Excel: Stock_Leaderboard ──
    ws = wb.create_sheet("Stock_Leaderboard")
    if slb:
        slb_cols = [k for k in slb[0] if k != "_composite"]
        _xl_df(ws, pd.DataFrame(slb)[slb_cols], sort_by=["PF", "OppScore"])
    _xl_hdr(ws); _xl_autow(ws)

    # ── Excel: Daily_Rankings ──
    ws = wb.create_sheet("Daily_Rankings")
    rank_cols = [c for c in ["ticker", "date", "daily_rank", "opportunity_score",
                               "liquidity_score", "orderflow_score", "volatility_score",
                               "quotes_score", "options_score", "cross_market_score"]
                  if c in daily_df.columns]
    _xl_df(ws, _strip_tz(daily_df.head(2000).sort_values("opportunity_score", ascending=False))[rank_cols])
    _xl_hdr(ws); _xl_autow(ws)

    # ── Excel: Trades ──
    ws = wb.create_sheet("Trades")
    td_cols = ["ticker", "entry_time", "exit_time", "direction", "entry_price", "exit_price",
               "stop_price", "shares", "position_value", "margin_used", "risk_$",
               "pnl_%", "pnl_$", "commission_$", "running_equity", "result", "strategy", "rr_ratio"]
    td_out = _strip_tz(pd.DataFrame(all_trades[:5000]))
    td_avail = [c for c in td_cols if c in td_out.columns]
    if not td_out.empty:
        _xl_df(ws, td_out[td_avail])
        if "pnl_$" in td_avail:
            pnl_col = td_avail.index("pnl_$") + 1
            _xl_cfmt(ws, pnl_col)
    _xl_hdr(ws, nc=len(td_avail)); _xl_autow(ws)

    # ── Excel: Time_of_Day ──
    ws = wb.create_sheet("(C) Time_of_Day")
    tz_data = []
    for (s_name, rr, tz), pnls in sorted(time_zone_stats.items(), key=lambda x: -len(x[1])):
        if len(pnls) < 3: continue
        pnl_arr = np.array(pnls)
        wins = (pnl_arr > 0).sum(); n = len(pnl_arr)
        wr = wins / n * 100 if n > 0 else 0
        total_win = pnl_arr[pnl_arr > 0].sum(); total_loss = abs(pnl_arr[pnl_arr < 0].sum())
        pf = _safe_div(total_win, total_loss, 999)
        std = pnl_arr.std(); exp = pnl_arr.mean()
        sh = _safe_div(exp, std) * np.sqrt(252) if std > 0 else 0
        tz_data.append([s_name, f"1:{rr}", tz, round(exp, 4), n, round(wr, 1), round(pnl_arr.sum(), 4), round(pf, 2), round(sh, 2)])
    tz_data.sort(key=lambda r: r[7], reverse=True)  # sort by PF col
    tz_rows = [["Strategy", "R:R", "Time Zone", "Avg PnL %", "N Trades", "Win Rate", "Total PnL", "PF", "Sharpe"]] + tz_data
    for ri, row in enumerate(tz_rows, 1):
        for ci, v in enumerate(row, 1): ws.cell(row=ri, column=ci, value=v)
    _xl_hdr(ws, nc=len(tz_rows[0])); _xl_autow(ws)

    # Separate bar charts: PF, Win Rate, and Sharpe+PnL for clean scaling
    if len(tz_data) > 0:
        n_rows = len(tz_data) + 1
        cats = Reference(ws, min_col=3, min_row=2, max_row=n_rows)
        cw = 40; ch = 18

        ch1 = BarChart()
        ch1.title = "Profit Factor"; ch1.y_axis.title = "PF"
        ch1.style = 10; ch1.height = ch; ch1.width = cw
        ch1.add_data(Reference(ws, min_col=8, min_row=1, max_row=n_rows, max_col=8), titles_from_data=True)
        ch1.set_categories(cats)
        ch1.series[0].graphicalProperties.solidFill = "2F5496"
        ch1.legend = None
        ws.add_chart(ch1, f"L{n_rows + 3}")

        ch2 = BarChart()
        ch2.title = "Win Rate %"; ch2.y_axis.title = "Win Rate (%)"
        ch2.style = 10; ch2.height = ch; ch2.width = cw
        ch2.add_data(Reference(ws, min_col=6, min_row=1, max_row=n_rows, max_col=6), titles_from_data=True)
        ch2.set_categories(cats)
        ch2.series[0].graphicalProperties.solidFill = "4472C4"
        ch2.legend = None
        ws.add_chart(ch2, f"L{n_rows + ch + 5}")

    # ── Excel: Day_of_Week ──
    ws = wb.create_sheet("(C) Day_of_Week")
    dow_data = []
    if not trade_df.empty and "day_of_week" in trade_df.columns:
        for dow in ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]:
            d_t = trade_df[trade_df["day_of_week"] == dow]
            if len(d_t) < 3: continue
            m = _trade_metrics(d_t)
            dow_data.append([dow, m["n_trades"], round(m["win_rate"], 1), round(m["profit_factor"], 2),
                             round(m["expectancy"], 2), round(m["total_pnl"], 2),
                             round(m["sharpe"], 2), round(m["avg_win"], 2),
                             round(m["avg_loss"], 2), 0])
    dow_data.sort(key=lambda r: r[3], reverse=True)  # sort by PF
    dow_rows = [["Day", "N Trades", "Win Rate", "PF", "Expectancy$", "Total PnL$", "Sharpe", "AvgWin$", "AvgLoss$", "MaxDD%"]] + dow_data
    for ri, row in enumerate(dow_rows, 1):
        for ci, v in enumerate(row, 1): ws.cell(row=ri, column=ci, value=v)
    _xl_hdr(ws, nc=len(dow_rows[0])); _xl_autow(ws)

    # Separate bar charts
    if len(dow_data) > 0:
        n_rows = len(dow_data) + 1
        cats = Reference(ws, min_col=1, min_row=2, max_row=n_rows)
        cw = 36; ch = 16

        ch1 = BarChart()
        ch1.title = "Profit Factor"; ch1.y_axis.title = "PF"
        ch1.style = 10; ch1.height = ch; ch1.width = cw
        ch1.add_data(Reference(ws, min_col=4, min_row=1, max_row=n_rows, max_col=4), titles_from_data=True)
        ch1.set_categories(cats)
        ch1.series[0].graphicalProperties.solidFill = "2F5496"
        ch1.legend = None
        ws.add_chart(ch1, f"A{n_rows + 3}")

        ch2 = BarChart()
        ch2.title = "Win Rate %"; ch2.y_axis.title = "Win Rate (%)"
        ch2.style = 10; ch2.height = ch; ch2.width = cw
        ch2.add_data(Reference(ws, min_col=3, min_row=1, max_row=n_rows, max_col=3), titles_from_data=True)
        ch2.set_categories(cats)
        ch2.series[0].graphicalProperties.solidFill = "4472C4"
        ch2.legend = None
        ws.add_chart(ch2, f"A{n_rows + ch + 5}")

    # ── Excel: Factor_Analysis ──
    ws = wb.create_sheet("Factor_Analysis")
    fa_scores = ["liquidity_score", "orderflow_score", "volatility_score", "quotes_score", "options_score", "cross_market_score"]
    fa_header = ["Factor", "Quintile", "N Trades", "Win Rate", "PF", "Expectancy$", "Total PnL$", "Sharpe"]
    fa_data = []
    if not trade_df.empty and not daily_df.empty:
        for score_name in fa_scores:
            if score_name not in daily_df.columns: continue
            try:
                daily_df["_q_temp"] = pd.qcut(daily_df[score_name].fillna(50), 5, labels=["Q1","Q2","Q3","Q4","Q5"], duplicates="drop")
                for q in sorted(daily_df["_q_temp"].dropna().unique()):
                    q_d = daily_df[daily_df["_q_temp"] == q][["ticker", "date"]]
                    q_t = trade_df.merge(q_d, on="ticker", how="inner") if not trade_df.empty else pd.DataFrame()
                    m = _trade_metrics(q_t)
                    fa_data.append([score_name.replace("_score",""), str(q), m["n_trades"], round(m["win_rate"],1),
                                    round(m["profit_factor"],2), round(m["expectancy"],2),
                                    round(m["total_pnl"],2), round(m["sharpe"],2)])
            except Exception: pass
    fa_data.sort(key=lambda r: (r[1], r[4]), reverse=False)  # sort by quintile then PF
    for ri, row in enumerate([fa_header] + fa_data, 1):
        for ci, v in enumerate(row, 1): ws.cell(row=ri, column=ci, value=v)
    _xl_hdr(ws, nc=len(fa_header)); _xl_autow(ws)

    # ── Excel: Options_Analysis ──
    ws = wb.create_sheet("Options_Analysis")
    opt_cols = ["opt_straddle_pct", "opt_expected_move_pct", "opt_total_vol", "opt_put_call_ratio", "chain_atm_iv", "chain_iv_skew", "options_score"]
    opt_header = ["Metric", "Quintile", "N Trades", "Win Rate", "PF", "Expectancy$", "Total PnL$", "Sharpe"]
    opt_data = []
    if not trade_df.empty and not daily_df.empty:
        for col in opt_cols:
            if col not in daily_df.columns: continue
            try:
                vals = daily_df[col].replace([np.inf, -np.inf], np.nan).fillna(0)
                daily_df["_oq"] = pd.qcut(vals, 5, labels=["Q1","Q2","Q3","Q4","Q5"], duplicates="drop")
                for q in sorted(daily_df["_oq"].dropna().unique()):
                    q_d = daily_df[daily_df["_oq"] == q][["ticker", "date"]]
                    q_t = trade_df.merge(q_d, on="ticker", how="inner") if not trade_df.empty else pd.DataFrame()
                    m = _trade_metrics(q_t)
                    opt_data.append([col, str(q), m["n_trades"], round(m["win_rate"],1),
                                     round(m["profit_factor"],2), round(m["expectancy"],2),
                                     round(m["total_pnl"],2), round(m["sharpe"],2)])
            except Exception: pass
    opt_data.sort(key=lambda r: (r[1], r[4]), reverse=False)  # sort by quintile then PF
    for ri, row in enumerate([opt_header] + opt_data, 1):
        for ci, v in enumerate(row, 1): ws.cell(row=ri, column=ci, value=v)
    _xl_hdr(ws, nc=len(opt_header)); _xl_autow(ws)

    # ── Excel: Microstructure ──
    ws = wb.create_sheet("Microstructure")
    ms_cols = ["avg_spread_qu", "qu_imbalance_mean", "tr_delta_sum", "tr_aggression_ratio", "trade_freq"]
    ms_header = ["Metric", "Quintile", "N Trades", "Win Rate", "PF", "Expectancy$", "Total PnL$", "Sharpe"]
    ms_data = []
    if not trade_df.empty and not daily_df.empty:
        for col in ms_cols:
            if col not in daily_df.columns: continue
            try:
                vals = daily_df[col].replace([np.inf, -np.inf], np.nan).fillna(0)
                daily_df["_mq"] = pd.qcut(vals, 5, labels=["Q1","Q2","Q3","Q4","Q5"], duplicates="drop")
                for q in sorted(daily_df["_mq"].dropna().unique()):
                    q_d = daily_df[daily_df["_mq"] == q][["ticker", "date"]]
                    q_t = trade_df.merge(q_d, on="ticker", how="inner") if not trade_df.empty else pd.DataFrame()
                    m = _trade_metrics(q_t)
                    ms_data.append([col, str(q), m["n_trades"], round(m["win_rate"],1),
                                    round(m["profit_factor"],2), round(m["expectancy"],2),
                                    round(m["total_pnl"],2), round(m["sharpe"],2)])
            except Exception: pass
    ms_data.sort(key=lambda r: (r[1], r[4]), reverse=False)
    for ri, row in enumerate([ms_header] + ms_data, 1):
        for ci, v in enumerate(row, 1): ws.cell(row=ri, column=ci, value=v)
    _xl_hdr(ws, nc=len(ms_header)); _xl_autow(ws)

    # ── Excel: Equity_Curve (trade log) ──
    ws = wb.create_sheet("Equity_Curve")
    if not trade_df.empty:
        trade_df_sorted = trade_df.sort_values("exit_time_dt") if "exit_time_dt" in trade_df.columns else trade_df.copy()
        eq_rows = [["Trade#", "Exit Time", "Strategy", "R:R", "P&L$", "Cumulative P&L$", "Running Equity"]]
        cum = 0; eq_start = capital.starting_cash
        for i, (_, t) in enumerate(trade_df_sorted.iterrows(), 1):
            cum += t.get("pnl_$", 0)
            eq_rows.append([i, str(t.get("exit_time", "")), t.get("strategy", ""),
                            f"1:{t.get('rr_ratio','')}",
                            round(t.get("pnl_$", 0), 2), round(cum, 2), round(eq_start + cum, 2)])
        if not port_df.empty:
            eq_rows.append(["", "", "", "", "", "", ""])
            eq_rows.append(["Portfolio Summary", "", "Strategy", "Total Return$", "Return%", "MaxDD%", "Trades"])
            for _, pr in port_df.iterrows():
                eq_rows.append(["", "", pr.get("strategy",""), round(pr.get("total_return_$",0),2),
                                round(pr.get("total_return_%",0),2), round(pr.get("max_drawdown_%",0),2),
                                pr.get("total_trades",0)])
        for ri, row in enumerate(eq_rows, 1):
            for ci, v in enumerate(row, 1): ws.cell(row=ri, column=ci, value=v)
        _xl_hdr(ws, nc=len(eq_rows[0])); _xl_autow(ws)
        if len(eq_rows) > 1:
            _xl_cfmt(ws, 5)
    else:
        ws.cell(row=1, column=1, value="No trades executed")

    # ── Excel: Equity_Curves (one line per Strategy×R:R) ──
    ws_eqcurves = wb.create_sheet("(C) Equity_Curves")
    if not trade_df.empty and "pnl_$" in trade_df.columns:
        trade_df_sorted = trade_df.sort_values("exit_time_dt") if "exit_time_dt" in trade_df.columns else trade_df.copy()
        combos = trade_df_sorted.groupby(["strategy", "rr_ratio"])
        combo_keys = []
        for (s, rr), grp in combos:
            if len(grp) < 2: continue
            combo_keys.append((s, rr))

        if combo_keys:
            # Shorten strategy names for legend
            _short = {
                "Momentum Breakout": "MomBrk", "VWAP Reversion": "VWAP",
                "RSI Scalp": "RSI", "EMA Pullback": "EMA",
                "Opening Range Breakout": "ORB",
                "Delta Confirmation Breakout": "DeltaBrk",
                "Quote Imbalance Breakout": "QuoteBrk",
                "VWAP + Delta": "VWAP+Del",
                "Liquidity Vacuum": "LiqVac",
                "VWAP + Momentum Confluence": "VWAP+Mom",
                "VWAP + Quote Imbalance Confluence": "VWAP+Qu",
                "VWAP + Delta Confluence": "VWAP+DelC",
                "VWAP Triple Confluence": "VWAPx3",
                "RSI + Momentum Confluence": "RSI+Mom",
                "RSI + Delta Confluence": "RSI+Del",
                "EMA + Momentum Confluence": "EMA+Mom",
                "EMA + Delta Confluence": "EMA+Del",
                "ORB + Volume Confluence": "ORB+Vol",
                "Momentum + Quote Confluence": "Mom+Qu",
                "Momentum + Delta + Quote Confluence": "MomDQu",
            }
            max_len = max(
                len(trade_df_sorted[(trade_df_sorted["strategy"] == s) & (trade_df_sorted["rr_ratio"] == rr)])
                for s, rr in combo_keys
            )
            ec_header = ["Trade#"] + [f"{_short.get(s, s[:8])} 1:{rr}" for s, rr in combo_keys]
            ec_rows = []
            for i in range(max_len):
                row = [i + 1]
                for s, rr in combo_keys:
                    subset = trade_df_sorted[(trade_df_sorted["strategy"] == s) & (trade_df_sorted["rr_ratio"] == rr)]
                    pnl_values = subset["pnl_$"].values
                    if i < len(pnl_values):
                        cum_pnl = pnl_values[:i + 1].sum()
                        row.append(round(cum_pnl, 2))
                    else:
                        row.append(None)
                ec_rows.append(row)
            for ci, h in enumerate(ec_header, 1):
                ws_eqcurves.cell(row=1, column=ci, value=h)
            for ri, row in enumerate(ec_rows, 2):
                for ci, v in enumerate(row, 1):
                    ws_eqcurves.cell(row=ri, column=ci, value=v)
            _xl_hdr(ws_eqcurves, nc=len(ec_header)); _xl_autow(ws_eqcurves)

            # Line chart: wider, legend on bottom
            if len(combo_keys) > 0:
                chart = LineChart()
                chart.title = "Equity Curves by Strategy × R:R"
                chart.y_axis.title = "Cumulative P&L ($)"
                chart.x_axis.title = "Trade #"
                chart.style = 10; chart.height = 20; chart.width = 52
                chart.legend.position = 'b'
                colors = ["2F5496", "C00000", "548235", "BF8F00", "7030A0",
                          "00B0F0", "FF6600", "4472C4", "A5A5A5", "FFC000"]
                for idx, (s, rr) in enumerate(combo_keys):
                    col_idx = idx + 2
                    data_ref = Reference(ws_eqcurves, min_col=col_idx, min_row=1,
                                         max_row=min(max_len + 1, 501), max_col=col_idx)
                    chart.add_data(data_ref, titles_from_data=True)
                    chart.series[idx].graphicalProperties.solidFill = colors[idx % len(colors)]
                cats_ref = Reference(ws_eqcurves, min_col=1, min_row=2, max_row=min(max_len + 1, 501))
                chart.set_categories(cats_ref)
                ws_eqcurves.add_chart(chart, f"A{max_len + 4}")
    else:
        ws_eqcurves.cell(row=1, column=1, value="No trades with $ metrics available")

    # ── Excel: Data_Quality ──
    ws = wb.create_sheet("Data_Quality")
    dq_header = ["Ticker", "OHLCV", "Trades", "Quotes", "Options", "Chains",
                "Missing%", "Backtested", "N Trades"]
    dq_data = []
    for ticker in sorted(availability.keys()):
        av = availability[ticker]
        present = sum(1 for v in av.values() if v)
        total = len(av)
        missing_pct = (1 - present / total) * 100 if total > 0 else 100
        bt = "Y" if ticker in backtest_tickers else ""
        tk_cnt = len(trade_df[trade_df["ticker"] == ticker]) if not trade_df.empty else 0
        dq_data.append([ticker, "Y" if av["OHLCV"] else "-", "Y" if av["Trades"] else "-",
                        "Y" if av["Quotes"] else "-", "Y" if av["Options"] else "-",
                        "Y" if av["Chains"] else "-", round(missing_pct, 1), bt, tk_cnt])
    dq_data.sort(key=lambda r: r[6])  # sort by missing_pct ascending
    for ri, row in enumerate([dq_header] + dq_data, 1):
        for ci, v in enumerate(row, 1): ws.cell(row=ri, column=ci, value=v)
    _xl_hdr(ws, nc=len(dq_header)); _xl_autow(ws)

    # ── Excel: Keep existing sheets for backward compat ──
    ws = wb.create_sheet("Strategy Signals")
    if not sum_df.empty: _xl_df(ws, sum_df, sort_by="profit_factor"); _xl_hdr(ws); _xl_autow(ws)
    ws = wb.create_sheet("Cross Market Confirm")
    if not daily_df.empty:
        xm_cols = [c for c in ["ticker", "date", "opportunity_score", "cross_market_score",
                                "orderflow_score", "quotes_score", "options_score"] if c in daily_df.columns]
        _xl_df(ws, _strip_tz(daily_df.sort_values("cross_market_score", ascending=False).head(100))[xm_cols])
        _xl_hdr(ws); _xl_autow(ws)
    ws = wb.create_sheet("Strategy Summary")
    if not sum_df.empty:
        sm_df = sum_df.groupby(["strategy", "rr_ratio"]).agg(
            n_tickers=("ticker","nunique"), total_trades=("total_trades","sum"),
            avg_wr=("win_rate_%","mean"), avg_pf=("profit_factor","mean"),
            total_pnl=("total_pnl_%","sum"), avg_exp=("expectancy_%","mean"),
            avg_sharpe=("sharpe_approx","mean")).reset_index()
        _xl_df(ws, sm_df, sort_by=["avg_pf", "total_pnl"])
        _xl_hdr(ws); _xl_autow(ws)
    ws = wb.create_sheet("(C) R-R Comparison")
    row_offset = 1
    for rr in rr_ratios:
        sub = sum_df[sum_df["rr_ratio"] == rr] if not sum_df.empty else pd.DataFrame()
        if sub.empty: continue
        agg = sub.groupby("strategy").agg(avg_wr=("win_rate_%","mean"), avg_pf=("profit_factor","mean"),
                                          total_pnl=("total_pnl_%","sum"), n=("ticker","count")).reset_index()
        ws.cell(row=row_offset, column=1, value=f"R:R 1:{rr}").font = Font(bold=True, size=12)
        row_offset += 1; _xl_hdr(ws, row=row_offset, nc=4); row_offset += 1
        for ri, row in enumerate(dataframe_to_rows(agg, index=False, header=True), row_offset):
            for ci, v in enumerate(row, 1): ws.cell(row=ri, column=ci, value=v)
        row_offset += len(agg) + 3
    _xl_autow(ws)

    # Summary bar chart at bottom of R-R sheet
    if not sum_df.empty and sum_df["rr_ratio"].nunique() > 0:
        rr_summary = sum_df.groupby("rr_ratio").agg(
            avg_pf=("profit_factor", "mean"), total_pnl=("total_pnl_%", "sum"),
            avg_wr=("win_rate_%", "mean")
        ).reset_index()
        chart_start = row_offset + 2
        rr_summary["rr_label"] = rr_summary["rr_ratio"].apply(lambda x: f"1:{x}")
        for i, (_, r) in enumerate(rr_summary.iterrows()):
            for j, col in enumerate(["rr_label", "avg_pf", "total_pnl", "avg_wr"]):
                ws.cell(row=chart_start + i, column=j + 1, value=r[col])
        ws.cell(row=chart_start, column=1, value="R:R Summary").font = Font(bold=True, size=11)
        n_rr = len(rr_summary)
        cats = Reference(ws, min_col=1, min_row=chart_start + 1, max_row=chart_start + n_rr)

        ch1 = BarChart()
        ch1.title = "Profit Factor by R:R"
        ch1.y_axis.title = "PF"; ch1.style = 10
        ch1.height = 18; ch1.width = 36
        ch1.add_data(Reference(ws, min_col=2, min_row=chart_start, max_row=chart_start + n_rr, max_col=2), titles_from_data=True)
        ch1.set_categories(cats)
        ch1.series[0].graphicalProperties.solidFill = "2F5496"
        ch1.legend = None
        ws.add_chart(ch1, f"F{chart_start}")

        ch2 = BarChart()
        ch2.title = "Win Rate % by R:R"
        ch2.y_axis.title = "Win Rate (%)"; ch2.style = 10
        ch2.height = 18; ch2.width = 36
        ch2.add_data(Reference(ws, min_col=4, min_row=chart_start, max_row=chart_start + n_rr, max_col=4), titles_from_data=True)
        ch2.set_categories(cats)
        ch2.series[0].graphicalProperties.solidFill = "4472C4"
        ch2.legend = None
        ws.add_chart(ch2, f"F{chart_start + 20}")
    ws = wb.create_sheet("Drawdown")
    if not port_df.empty:
        _xl_df(ws, port_df.groupby("strategy")["max_drawdown_%"].agg(["mean","max","min"]).reset_index(), sort_by="mean")
        _xl_hdr(ws); _xl_autow(ws)
    ws = wb.create_sheet("Performance")
    if not port_df.empty:
        _xl_df(ws, port_df.groupby(["strategy","rr_ratio"]).agg(
            total_return=("total_return_$","sum"), avg_return=("total_return_%","mean"),
            max_dd=("max_drawdown_%","mean"), n_trades=("total_trades","sum"),
            commission=("total_commission","sum")).reset_index(), sort_by="total_return")
        _xl_hdr(ws); _xl_autow(ws)

    # ── Excel: README (documentation sheet) ──
    readme_lines = [
        ["SCALPING ANALYSIS — EXCEL README"],
        [""],
        ["SHEET OVERVIEW"],
        [""],
        ["SUMMARY SHEETS (top of workbook)"],
        ["  Executive_Summary       Best strategy/R:R, data coverage, position sizing"],
        ["  (C) Strategy_Comparison  Ranked table + 4 charts: PF, WinRate, Sharpe+PnL%, Trades"],
        ["  Stock_Leaderboard        Per-ticker scores + best strategy + backtest metrics"],
        ["  Daily_Rankings           Daily top-stock table with component scores (0-100)"],
        [""],
        ["TRADE DETAIL SHEETS"],
        ["  Trades                   Individual trade log: entry/exit/stop, P&L, risk, margin"],
        ["  (C) Equity_Curve         Raw trade log with cumulative P&L"],
        ["  (C) Equity_Curves        One cumulative P&L column per Strategy×R:R combo + line chart"],
        [""],
        ["ANALYSIS SHEETS"],
        ["  (C) Time_of_Day          Performance by ET time bucket (09:30-09:45, etc.) + PF/WR charts"],
        ["  (C) Day_of_Week          Performance by weekday + PF/WR charts"],
        ["  Factor_Analysis          Performance by quintile for each composite score (Liq/OF/Vol/Qu/Opt/XMkt)"],
        ["  Options_Analysis         Performance by straddle%, expected move, IV, put/call quintiles"],
        ["  Microstructure           Performance by spread, quote imbalance, delta, aggression quintiles"],
        [""],
        ["COMPARISON SHEETS"],
        ["  (C) R-R Comparison       Aggregated PF & Win Rate by R:R ratio across all strategies + charts"],
        ["  Strategy Signals         Per-ticker raw strategy trade summaries (pre-portfolio)"],
        ["  Cross Market Confirm     Top 100 daily rows by cross-market confirmation score"],
        ["  Strategy Summary         Averaged per-strategy × R:R metrics (n_tickers, avg_wr, avg_pf, etc.)"],
        ["  Drawdown                 Max drawdown by strategy (mean/max/min)"],
        ["  Performance              Total return, avg return, max DD by strategy+R:R"],
        [""],
        ["REFERENCE SHEETS"],
        ["  Data_Quality             Per-ticker: which datasets exist, missing%, backtest status"],
        ["  README                   This sheet"],
        [""],
        ["CHART LEGEND"],
        ["  Dark blue bars (#2F5496)  = Profit Factor or Sharpe Ratio"],
        ["  Medium blue (#4472C4)     = Win Rate %"],
        ["  Gold/Yellow (#FFC000)     = Total PnL% (on cluster charts)"],
        ["  Purple (#7030A0)          = Total Trades"],
        ["  Equity_Curves uses distinct per-strategy line colours"],
        [""],
        ["KEY METRICS — HOW TO READ THEM"],
        [""],
        ["  Profit Factor (PF)        Gross wins / gross losses. > 1 = profitable. "],
        ["                            Higher = better risk/reward. Below 1 = losing."],
        [""],
        ["  Win Rate (WR)             % of resolved trades that won. With R:R 2:1, ~40% WR"],
        ["                            is already profitable. WR alone is insufficient; must be"],
        ["                            paired with PF and expectancy."],
        [""],
        ["  Expectancy ($)            Average profit per trade (including losses). Positive = "],
        ["                            system has edge. Magnitude matters less than consistency."],
        [""],
        ["  Sharpe Ratio              Risk-adjusted return. > 1.0 = good. Compares expectancy to"],
        ["                            volatility of returns. Higher = smoother equity curve."],
        [""],
        ["  Max Drawdown %            Largest peak-to-trough decline. Lower is better. "],
        ["                            > 20% suggests position sizing may be too aggressive."],
        [""],
        ["  Total PnL%                Sum of all trade percentage P&L (includes timeouts)."],
        ["                            Negative with high WR means timeouts are dragging results."],
        [""],
        ["  Opportunity Score (0-100) Composite ranking combining liquidity, order flow, "],
        ["                            volatility, quotes, options, and cross-market confirmation."],
        ["                            Higher = better scalping candidate for that day."],
        [""],
        ["IMPORTANT CAVEATS"],
        [""],
        ["  - Ranks use ONLY information available at that timestamp (no look-ahead bias)"],
        ["  - margin_used ≠ risk_$: margin = position_value × margin_rate; risk_$ = actual loss if stopped"],
        ["  - Timeout trades are included in Total PnL% but excluded from Win Rate"],
        ["  - <20 trades = LOW sample (flagged); interpret with caution"],
        ["  - Confluence strategies filter primary signals through secondary confirmations"],
        ["    (e.g., VWAP + Momentum Confluence = VWAP-signaled trades where Momentum"],
        ["     also signals same direction at the same bar)"],
        ["  - All charts use separate axes per metric — no mixed-scale distortion"],
        [""],
        ["SHEETS MARKED '(C) ' contain embedded Excel charts."],
        [""],
        ["Generated by scripts/strategy/scalping/scalping_analysis.py"],
    ]

    ws_readme = wb.create_sheet("README")
    for ri, row in enumerate(readme_lines, 1):
        for ci, v in enumerate(row, 1):
            cell = ws_readme.cell(row=ri, column=ci, value=v)
            if ri == 1:
                cell.font = Font(bold=True, size=14, color="2F5496")
            elif str(v).startswith("  ") and not str(v).startswith("    "):
                cell.font = Font(bold=True, size=11)
            elif str(v).startswith("SUMMARY") or str(v).startswith("TRADE") or str(v).startswith("ANALYSIS") or str(v).startswith("COMPARISON") or str(v).startswith("REFERENCE") or str(v).startswith("CHART") or str(v).startswith("KEY") or str(v).startswith("IMPORTANT") or str(v).startswith("SHEETS"):
                cell.font = Font(bold=True, size=12, color="2F5496")
    ws_readme.column_dimensions["A"].width = 90
    ws_readme.freeze_panes = "A2"

    wb.save(output_path)
    log.info("Excel saved to: %s", output_path)

    # ═══════════════════════════════════════════════════════════════════
    # KEY FINDINGS
    # ═══════════════════════════════════════════════════════════════════

    print("=" * W)
    print("  KEY FINDINGS")
    print("=" * W)
    best_strat = lb[0] if lb else {}
    best_ticker = slb[0] if slb else {}
    print(f"  1.  Best strategy:      {best_strat.get('Strategy','N/A')}")
    print(f"  2.  Best R:R:           {best_strat.get('R:R','N/A')}")
    print(f"  3.  Best ticker:        {best_ticker.get('Ticker','N/A')} (score={best_ticker.get('OppScore',0):.0f})")
    print(f"  4.  Best opp-score:     {best_ticker.get('OppScore',0):.0f}")
    print(f"  5.  Best time of day:   {max(time_zone_stats.items(), key=lambda x: len(x[1]))[0][2] if time_zone_stats else 'N/A'}")

    # Aggregate PF across all trades
    if not trade_df.empty:
        agg_m = _trade_metrics(trade_df)
        print(f"  6.  Best factor combo:  Vol={best_ticker.get('Vol',0):.0f}, OF={best_ticker.get('OF',0):.0f}, Liq={best_ticker.get('Liq',0):.0f}")
        print(f"  7.  Best options regime: score={best_ticker.get('Opt',0):.0f}")
        print(f"  8.  Best micro regime:   quotes_score={best_ticker.get('Qu',0):.0f}")
        print(f"  9.  Overall PF:         {agg_m['profit_factor']:.2f}")
        print(f"  10. Overall expectancy:  {_fmt_dollar(agg_m['expectancy'])}")
        print(f"  11. Overall P&L:        {_fmt_dollar(agg_m['total_pnl'])}")
        print(f"  12. Max drawdown:       {_fmt_pct(best_strat.get('MaxDD%',0))}")
    else:
        for i in range(6, 13):
            print(f"  {i}. {'(no trades)' if i == 6 else ''}")

    # Overall performance recap
    print(f"\n  Total time: {mins}m {secs}s  |  "
          f"Data: OHLCV={counts['OHLCV']} Trades={counts['Trades']} "
          f"Quotes={counts['Quotes']} Options={counts['Options']} Chains={counts['Chains']}")
    print(f"  Position sizing: risk/trade={capital.risk_pct*100:.1f}% "
          f"= ${capital._risk_dollars():,.0f}  |  "
          f"Output: {output_path}")
    print("=" * W)


def main():
    parser = argparse.ArgumentParser(
        description="Multi-Factor Daily Scalping Scanner + Backtester",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Position sizing note:\n"
            "  --risk-amount N    risk N%% of starting-cash per trade (the actual dollars you can lose).\n"
            "  margin_used is NOT the same as risk — it's the buying-power requirement.\n"
            "  Use --max-position-pct to cap position size as %% of equity if you want to\n"
            "  limit concurrent margin consumption (lower = fewer shares per trade)."
        ),
    )
    parser.add_argument("--year", default="2025")
    parser.add_argument("--aggregate", default=DEFAULT_AGGREGATE,
                        choices=["1sec", "1min", "5min", "15min", "1H", "4H", "1D"],
                        help="Bar aggregate to load (default: 1sec). Data files must exist "
                             "under data/SPY/<agg>/<year>/ etc.")
    parser.add_argument("--top_n", type=int, default=40)
    parser.add_argument("--backtest_n", type=int, default=10)
    parser.add_argument("--num_trades", type=int, default=None,
                        help="Cap the number of trades backtested per strategy×R:R "
                             "(default: None = use all trades).")
    parser.add_argument("--date_range", type=str, default=None,
                        help="Restrict the analysis to a date window, YYYYMMDD-YYYYMMDD "
                             "(e.g. 20250101-20250331). Default: the whole --year.")
    parser.add_argument("--starting-cash", type=float, default=100_000.0)
    parser.add_argument("--risk-amount", type=float, default=2.0,
                        help="Risk per trade as %% of starting cash (default: 2%%). "
                             "At $100k starting cash, --risk-amount 1 = risk $1,000 per trade.")
    parser.add_argument("--output", default="data/scalping_multi_2025.xlsx")
    parser.add_argument("--max-position-pct", type=float, default=200.0,
                        help="Max position value as %% of equity before trade is SKIPPED "
                             "(default: 200%% = max position 2x equity). "
                             "Set lower to limit margin consumption per trade.")
    parser.add_argument("--compounding", action="store_true", default=False)
    parser.add_argument("--min-confirmations", type=int, default=2,
                        help="Min cross-market confirmations required")
    parser.add_argument("--rr", type=str, default=None,
                        help="Comma-separated R:R ratios to test (default: 1.5,2.0,2.5). "
                             "Example: --rr 1.5,2.0,2.5,3.0")
    parser.add_argument("--strategies", type=str, default=None,
                        help="Comma-separated strategies to test (default: all). "
                             "Choices: Momentum Breakout, VWAP Reversion, RSI Scalp, "
                             "EMA Pullback, Opening Range Breakout, "
                             "Delta Confirmation Breakout, Quote Imbalance Breakout, "
                             "VWAP + Delta, Liquidity Vacuum")
    parser.add_argument("--nprocs", type=int, default=1,
                        help="Number of parallel processes (default: 1). "
                             "Set to a higher value (e.g. 6) to parallelize feature "
                             "computation and backtesting across tickers.")
    args = parser.parse_args()

    rr_ratios = RR_RATIOS
    if args.rr:
        rr_ratios = [float(x.strip()) for x in args.rr.split(",")]

    enabled = None
    if args.strategies:
        raw = [x.strip() for x in args.strategies.split(",")]
        enabled = []
        for s in raw:
            # Partial match
            matched = [k for k in STRATEGIES if s.lower() in k.lower()]
            if matched:
                enabled.extend(matched)
            else:
                log.warning("Unknown strategy: %s", s)
        enabled = list(dict.fromkeys(enabled))  # dedupe, keep order
        if not enabled:
            enabled = None
            log.warning("No valid strategies matched; using all.")

    capital = CapitalManager(
        starting_cash=args.starting_cash,
        risk_pct=args.risk_amount / 100.0,
        compounding=args.compounding,
        max_position_pct=args.max_position_pct / 100.0,
    )

    date_range = _parse_date_range(args.date_range)

    os.makedirs(Path(args.output).parent, exist_ok=True)
    run_pipeline(args.year, args.top_n, args.backtest_n, args.num_trades,
                 capital, args.output, enabled, rr_ratios, nprocs=args.nprocs,
                 aggregate=args.aggregate, date_range=date_range)


if __name__ == "__main__":
    main()
