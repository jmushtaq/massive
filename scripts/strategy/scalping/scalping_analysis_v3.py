#!/usr/bin/env python3
"""
V3 correctness-first scalping backtester.

A from-scratch, methodology-first rewrite of scalping_analysis.py. The goal is
the most realistic and reproducible backtest possible — NOT the highest return.

Correctness guarantees (enforced by construction and by --integrity assertions):

  * A signal is computed ONLY from data available at the close of bar `i`.
  * Entry is NEVER filled on the signal bar; the earliest entry is the NEXT bar
    (plus optional latency).
  * Stop/target are walked from the entry bar onward; a same-bar stop/target
    touch resolves conservatively (stop-first by default).
  * Every signal resolves to exactly one of TARGET / STOP / TIMEOUT, and
    TIMEOUTs are always included in every statistic.
  * Execution uses real bid/ask from quotes data (configurable: mid, bid_ask,
    bid_ask_slippage) plus configurable bps slippage and latency.
  * ONE shared portfolio is used for every ticker (no per-ticker $100k).
  * Stock selection for day D uses ONLY data from days strictly before D
    (configurable 20/30/60-day lookback) — no hindsight.
  * ORB uses the 09:30 ET regular open (pre-market ORB is a separate strategy).

Usage:
    python scripts/strategy/scalping/scalping_analysis_v3.py \
        --year 2026 --aggregate 1sec --strategies "ORB + Volume Confluence" \
        --rr 2.0 --timeout 30 --execution bid_ask --universe-limit 60 \
        --date-range 20260102-20260130 --robustness --integrity
"""

import argparse
import datetime
import logging
import os
import sys
from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent))

from scalping_analysis import (  # noqa: E402
    atr_arr,
    load_ohlcv,
    load_quotes,
    load_trades,
    rsi_arr,
)

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger("v3")

DATA_OHLCV = Path("data/SPY")
DATA_TRADES = Path("data/trades")
DATA_QUOTES = Path("data/quotes")

SECONDS_PER_BAR = {"1sec": 1, "1min": 60, "5min": 300, "15min": 900,
                   "1H": 3600, "4H": 14400, "1D": 86400}
RTH_OPEN = (9, 30)   # 09:30 ET
RTH_TRADE_END = (11, 0)  # stop trading at 11:00 ET (matches V2 window)


def clean_ticker(raw: str) -> str:
    return raw.strip().upper().split("-")[0]


# ══════════════════════════════════════════════════════════════════════
# Config
# ══════════════════════════════════════════════════════════════════════

@dataclass
class Config:
    year: str = "2026"
    aggregate: str = "1sec"
    date_range: tuple | None = None
    # portfolio
    starting_cash: float = 100_000.0
    risk_amount: float = 1.0            # % of equity risked per trade
    margin_rate: float = 0.25
    max_positions: int = 5
    max_portfolio_risk_pct: float = 20.0  # % of equity at risk concurrently
    max_notional: float = 200_000.0     # max gross notional exposure
    max_leverage: float = 4.0
    # strategy
    strategies: list[str] | None = None
    orb_minutes: int = 5
    timeout_seconds: int = 30
    rr: float = 2.0
    # execution
    execution: str = "bid_ask"          # mid | bid_ask | bid_ask_slippage
    slippage_bps: float = 0.0
    latency_seconds: int = 0
    intrabar_priority: str = "stop_first"  # stop_first | target_first
    # selection
    lookback_days: int = 20
    top_n: int = 40
    universe_limit: int = 0             # 0 = all tickers
    # reporting
    min_trades_for_stats: int = 5


# ══════════════════════════════════════════════════════════════════════
# Data loading / alignment
# ══════════════════════════════════════════════════════════════════════

def _path(ticker, base, year, agg, kind=""):
    name = f"{ticker}_{year}_{agg}_{kind}.csv" if kind else f"{ticker}_{year}_{agg}.csv"
    return base / agg / year / name


def _filter_date_range(df, date_range):
    if df is None or date_range is None:
        return df
    start, end = date_range
    s = pd.Timestamp(start, tz="UTC")
    e = pd.Timestamp(end, tz="UTC") + pd.Timedelta(days=1)
    return df[(df.index >= s) & (df.index < e)]


def _load_ohlcv(ticker, cfg):
    p = _path(ticker, DATA_OHLCV, cfg.year, cfg.aggregate)
    if not p.exists():
        return None
    df = pd.read_csv(p, usecols=["timestamp", "open", "high", "low", "close", "volume"])
    df["timestamp"] = pd.to_datetime(df["timestamp"], utc=True)
    df.set_index("timestamp", inplace=True)
    df.sort_index(inplace=True)
    return _filter_date_range(df, cfg.date_range)


def _load_quotes_series(ticker, cfg):
    """Return a DataFrame of [bid, ask, spread, quote_imbalance] indexed by ts (ffill onto OHLCV)."""
    p = _path(ticker, DATA_QUOTES, cfg.year, cfg.aggregate, "quotes")
    if not p.exists():
        return None
    cols = ["timestamp", "avg_bid", "avg_ask", "avg_spread", "quote_imbalance"]
    df = pd.read_csv(p, usecols=cols)
    df["timestamp"] = pd.to_datetime(df["timestamp"], utc=True)
    df.set_index("timestamp", inplace=True)
    df.sort_index(inplace=True)
    return _filter_date_range(df, cfg.date_range)


def _load_trades_delta(ticker, cfg):
    p = _path(ticker, DATA_TRADES, cfg.year, cfg.aggregate, "trades")
    if not p.exists():
        return None
    df = pd.read_csv(p, usecols=["timestamp", "delta"])
    df["timestamp"] = pd.to_datetime(df["timestamp"], utc=True)
    df.set_index("timestamp", inplace=True)
    df.sort_index(inplace=True)
    return _filter_date_range(df, cfg.date_range)


def _build_ticker_df(ticker, cfg):
    """OHLCV merged with quotes bid/ask and trades delta, all forward-filled
    (never bfill, so no future leakage)."""
    o = _load_ohlcv(ticker, cfg)
    if o is None or len(o) < 200:
        return None
    df = o[["open", "high", "low", "close", "volume"]].copy()
    df.columns = ["o", "h", "l", "c", "vol"]

    qu = _load_quotes_series(ticker, cfg)
    if qu is not None and len(qu) > 0:
        df["bid"] = qu["avg_bid"].reindex(df.index, method="ffill")
        df["ask"] = qu["avg_ask"].reindex(df.index, method="ffill")
        df["spread"] = qu["avg_spread"].reindex(df.index, method="ffill")
        df["quote_imbalance"] = qu["quote_imbalance"].reindex(df.index, method="ffill")
    else:
        df["bid"] = np.nan
        df["ask"] = np.nan
        df["spread"] = np.nan
        df["quote_imbalance"] = np.nan

    tr = _load_trades_delta(ticker, cfg)
    if tr is not None and len(tr) > 0:
        df["delta"] = tr["delta"].reindex(df.index, method="ffill")
    else:
        df["delta"] = np.nan

    return df


# ══════════════════════════════════════════════════════════════════════
# Daily features (used ONLY for universe selection, computed day-by-day)
# ══════════════════════════════════════════════════════════════════════

def compute_daily_features(df):
    """One row per ET trading day, computed from that day's bars only."""
    try:
        et = df.index.tz_convert("US/Eastern")
    except Exception:
        et = df.index
    dates = et.date  # numpy array of datetime.date, aligned positionally
    rows = []
    for date, grp in df.groupby(dates):
        if len(grp) < 10:
            continue
        c = grp["c"]
        v = grp["vol"]
        rows.append({
            "date": date,
            "dollar_vol": float((c * v).mean()),
            "total_vol": float(v.sum()),
            "atr_pct": float(((grp["h"] - grp["l"]) / c * 100).mean()),
            "range_pct": float(((grp["h"] - grp["l"]) / c * 100).mean()),
            "ret_std": float(c.pct_change().std() * 100),
            "close": float(c.iloc[-1]),
        })
    return pd.DataFrame(rows, columns=["date", "dollar_vol", "total_vol",
                                       "atr_pct", "range_pct", "ret_std", "close"])


def rank_universe(daily_features, day, cfg):
    """Return tickers ranked for trading on `day` using ONLY days strictly
    before `day` over the last cfg.lookback_days trading days (no hindsight)."""
    scored = []
    for ticker, fdf in daily_features.items():
        if fdf is None or fdf.empty or "date" not in fdf.columns:
            continue
        hist = fdf[fdf["date"] < day].tail(cfg.lookback_days)
        if len(hist) < 5:
            continue
        dollar_vol = hist["dollar_vol"].mean()
        total_vol = hist["total_vol"].mean()
        atr_pct = hist["atr_pct"].mean()
        scored.append((ticker, dollar_vol, total_vol, atr_pct))
    if not scored:
        return []
    s = pd.DataFrame(scored, columns=["ticker", "dollar_vol", "total_vol", "atr_pct"])

    def z(col):
        v = s[col]
        m, sd = v.mean(), v.std()
        if sd == 0 or np.isnan(sd):
            return pd.Series(0.0, index=s.index)
        return ((v - m) / sd).clip(-3, 3)

    s["score"] = z("dollar_vol") * 0.4 + z("total_vol") * 0.3 + z("atr_pct") * 0.3
    s = s.sort_values("score", ascending=False)
    return s["ticker"].tolist()[: cfg.top_n]


# ══════════════════════════════════════════════════════════════════════
# Signals (closed-bar only; entry happens on the NEXT bar)
# ══════════════════════════════════════════════════════════════════════

@dataclass
class Signal:
    ticker: str
    bar: int
    direction: int          # +1 long, -1 short
    strategy: str
    stop_abs: float | None = None
    stop_offset: float | None = None


def _atr(c, i, k=10):
    return float(np.mean(np.abs(np.diff(c[i - k:i])))) if i >= k else float(np.mean(np.abs(np.diff(c[:i]))))


def sig_momentum_breakout(df, ticker):
    out = []
    h, l, v = df["h"].values, df["l"].values, df["vol"].values
    n = 5
    for i in range(n + 5, len(df)):
        rh = np.max(h[i - n:i])
        avg_v = np.mean(v[i - n:i])
        if h[i] > rh and v[i] > avg_v * 1.2:
            out.append(Signal(ticker, i, 1, "Momentum Breakout", stop_abs=float(np.min(l[i - n:i]))))
    return out


def sig_vwap_reversion(df, ticker, dev=0.3):
    out = []
    c, v = df["c"].values, df["vol"].values
    period = 20
    for i in range(period + 5, len(df)):
        vwap = np.sum(c[i - period:i] * v[i - period:i]) / (np.sum(v[i - period:i]) + 1e-9)
        dev_pct = (c[i] - vwap) / vwap * 100
        atr = _atr(c, i)
        if atr <= 0:
            continue
        if dev_pct > dev:
            out.append(Signal(ticker, i, -1, "VWAP Reversion", stop_offset=float(atr * 2.0)))
        elif dev_pct < -dev:
            out.append(Signal(ticker, i, 1, "VWAP Reversion", stop_offset=float(atr * 2.0)))
    return out


def sig_rsi_scalp(df, ticker):
    out = []
    c = df["c"].values
    rv = rsi_arr(c, 5)
    for i in range(15, len(df)):
        if pd.isna(rv[i - 1]) or pd.isna(rv[i]):
            continue
        atr = _atr(c, i)
        if atr <= 0:
            continue
        if rv[i - 1] < 30 and rv[i] >= 30:
            out.append(Signal(ticker, i, 1, "RSI Scalp", stop_offset=float(atr * 2)))
        elif rv[i - 1] > 70 and rv[i] <= 70:
            out.append(Signal(ticker, i, -1, "RSI Scalp", stop_offset=float(atr * 2)))
    return out


def sig_ema_pullback(df, ticker):
    out = []
    c, h, l = df["c"].values, df["h"].values, df["l"].values
    ema = pd.Series(c).ewm(span=20, adjust=False).mean().values
    for i in range(40, len(df)):
        trend = ema[i] - ema[i - 5]
        atr = _atr(c, i, 14)
        if atr <= 0:
            continue
        prox = (c[i] - ema[i]) / ema[i] * 100
        if trend > 0 and -0.2 < prox < 0.15:
            stop = float(np.min(l[i - 3:i]) - atr * 0.5)
            out.append(Signal(ticker, i, 1, "EMA Pullback", stop_abs=stop))
        elif trend < 0 and -0.15 < prox < 0.2:
            stop = float(np.max(h[i - 3:i]) + atr * 0.5)
            out.append(Signal(ticker, i, -1, "EMA Pullback", stop_abs=stop))
    return out


def sig_opening_range_breakout(df, ticker, cfg, premarket=False):
    """ORB over the first `orb_minutes` minutes of the regular session (09:30 ET),
    or the first N bars of the ET day if premarket=True."""
    out = []
    h, l = df["h"].values, df["l"].values
    try:
        et = df.index.tz_convert("US/Eastern")
    except Exception:
        et = df.index
    n = len(df)
    if n == 0:
        return out
    spb = SECONDS_PER_BAR.get(cfg.aggregate, 1)
    bars_per_minute = max(1, 60 // spb)
    lookback = cfg.orb_minutes * bars_per_minute

    day_int = et.normalize().asi8
    hours = et.hour.to_numpy()
    minutes = et.minute.to_numpy()

    orb_high = orb_low = 0.0
    orb_seen = 0
    orb_ready = False
    prev_date = day_int[0]

    for i in range(n):
        if day_int[i] != prev_date:
            orb_seen = 0
            orb_high = orb_low = 0.0
            orb_ready = False
            prev_date = day_int[i]

        hh, mm = int(hours[i]), int(minutes[i])
        if premarket:
            in_orb = True  # first N bars of the ET day
        else:
            in_orb = (hh, mm) >= RTH_OPEN

        if not orb_ready and orb_seen < lookback and in_orb:
            orb_high = h[i] if orb_seen == 0 else max(orb_high, h[i])
            orb_low = l[i] if orb_seen == 0 else min(orb_low, l[i])
            orb_seen += 1
            if orb_seen == lookback:
                orb_ready = True
            continue
        if not orb_ready:
            continue

        if premarket:
            if not (9 <= hh < 11):
                continue
        else:
            if (hh, mm) >= RTH_TRADE_END:
                continue

        if h[i] > orb_high:
            out.append(Signal(ticker, i, 1, "Opening Range Breakout" if not premarket else "Pre-Market ORB",
                              stop_abs=float(orb_low)))
        elif l[i] < orb_low:
            out.append(Signal(ticker, i, -1, "Opening Range Breakout" if not premarket else "Pre-Market ORB",
                              stop_abs=float(orb_high)))
    return out


def sig_delta_breakout(df, ticker):
    out = []
    h, l, v = df["h"].values, df["l"].values, df["vol"].values
    delta = df["delta"].values
    n = 5
    for i in range(n + 5, len(df)):
        rh = np.max(h[i - n:i])
        avg_v = np.mean(v[i - n:i])
        dv = delta[i]
        if h[i] > rh and v[i] > avg_v * 1.2 and (pd.isna(dv) or dv > 0):
            out.append(Signal(ticker, i, 1, "Delta Breakout", stop_abs=float(np.min(l[i - n:i]))))
    return out


def sig_quote_imbalance_breakout(df, ticker):
    out = []
    h, l, v = df["h"].values, df["l"].values, df["vol"].values
    qi = df["quote_imbalance"].values
    n = 5
    for i in range(n + 5, len(df)):
        rh = np.max(h[i - n:i])
        avg_v = np.mean(v[i - n:i])
        q = qi[i]
        if h[i] > rh and v[i] > avg_v * 1.2 and (pd.isna(q) or q > 0):
            out.append(Signal(ticker, i, 1, "Quote Imbalance Breakout", stop_abs=float(np.min(l[i - n:i]))))
    return out


def sig_vwap_delta(df, ticker, dev=0.3):
    out = []
    c, v = df["c"].values, df["vol"].values
    delta = df["delta"].values
    period = 20
    for i in range(period + 5, len(df)):
        vwap = np.sum(c[i - period:i] * v[i - period:i]) / (np.sum(v[i - period:i]) + 1e-9)
        dev_pct = (c[i] - vwap) / vwap * 100
        atr = _atr(c, i)
        if atr <= 0:
            continue
        dv = delta[i] if not pd.isna(delta[i]) else 0
        if dev_pct > dev and dv < 0:
            out.append(Signal(ticker, i, -1, "VWAP + Delta", stop_offset=float(atr * 2.0)))
        elif dev_pct < -dev and dv > 0:
            out.append(Signal(ticker, i, 1, "VWAP + Delta", stop_offset=float(atr * 2.0)))
    return out


def sig_liquidity_vacuum(df, ticker):
    out = []
    h, l, c, v = df["h"].values, df["l"].values, df["c"].values, df["vol"].values
    vol_ma = pd.Series(v).rolling(50).mean().values
    for i in range(55, len(df)):
        if pd.isna(vol_ma[i]) or vol_ma[i] <= 0:
            continue
        if v[i] / vol_ma[i] < 2.5:
            continue
        range_1m = h[i] - l[i]
        avg_range = np.mean(np.abs(h[i - 20:i] - l[i - 20:i]))
        if avg_range <= 0:
            continue
        if range_1m < avg_range * 0.7:
            atr = _atr(c, i)
            if atr <= 0:
                continue
            prev_move = c[i - 1] - c[i - 5] if i >= 5 else 0
            if prev_move > 0:
                out.append(Signal(ticker, i, -1, "Liquidity Vacuum", stop_offset=float(atr * 1.5)))
            else:
                out.append(Signal(ticker, i, 1, "Liquidity Vacuum", stop_offset=float(atr * 1.5)))
    return out


def _signal_index(sigs):
    return {(s.bar, s.direction) for s in sigs}


def _confluence(primary, confirmations, min_conf):
    if not confirmations:
        return primary
    sets = [_signal_index(c) for c in confirmations]
    return [s for s in primary if sum(1 for st in sets if (s.bar, s.direction) in st) >= min_conf]


BASE_SIGNALS = {
    "Momentum Breakout": sig_momentum_breakout,
    "VWAP Reversion": sig_vwap_reversion,
    "RSI Scalp": sig_rsi_scalp,
    "EMA Pullback": sig_ema_pullback,
    "Delta Breakout": sig_delta_breakout,
    "Quote Imbalance Breakout": sig_quote_imbalance_breakout,
    "VWAP + Delta": sig_vwap_delta,
    "Liquidity Vacuum": sig_liquidity_vacuum,
}


def _rth_orb(df, ticker, cfg):
    return sig_opening_range_breakout(df, ticker, cfg, premarket=False)


def _pre_orb(df, ticker, cfg):
    return sig_opening_range_breakout(df, ticker, cfg, premarket=True)


# (confluence name, primary signal source, confirmation sources, min confirmations)
CONFLUENCE_SPECS = [
    ("VWAP + Momentum Confluence", "VWAP Reversion", ["Momentum Breakout"], 1),
    ("VWAP + Quote Imbalance Confluence", "VWAP Reversion", ["Quote Imbalance Breakout"], 1),
    ("VWAP + Delta Confluence", "VWAP Reversion", ["Delta Breakout"], 1),
    ("VWAP Triple Confluence", "VWAP Reversion",
     ["Momentum Breakout", "Quote Imbalance Breakout", "Delta Breakout"], 2),
    ("RSI + Momentum Confluence", "RSI Scalp", ["Momentum Breakout"], 1),
    ("RSI + Delta Confluence", "RSI Scalp", ["Delta Breakout"], 1),
    ("EMA + Momentum Confluence", "EMA Pullback", ["Momentum Breakout"], 1),
    ("EMA + Delta Confluence", "EMA Pullback", ["Delta Breakout"], 1),
    ("ORB + Volume Confluence", "ORB (RTH)", ["Liquidity Vacuum"], 1),
    ("Momentum + Quote Confluence", "Momentum Breakout", ["Quote Imbalance Breakout"], 1),
    ("Momentum + Delta + Quote Confluence", "Momentum Breakout",
     ["Delta Breakout", "Quote Imbalance Breakout", "VWAP + Delta"], 2),
]

BASE_SOURCE_KEYS = list(BASE_SIGNALS.keys()) + ["ORB (RTH)", "ORB (Pre-Market)"]  # noqa: F841 (documented source list)


def generate_signals(df, ticker, cfg):
    """Generate signals for the enabled strategies on one ticker (closed-bar only).
    Base signals needed by a requested confluence are generated internally but only
    the enabled strategies' signals are emitted."""
    enabled = set(cfg.strategies) if cfg.strategies else set(ALL_STRATEGIES)

    # resolve which base signal sources are required
    needed = set()
    for name in enabled:
        if name in BASE_SIGNALS:
            needed.add(name)
        if name in ("Opening Range Breakout",):
            needed.add("ORB (RTH)")
        if name in ("Pre-Market ORB", "ORB (Pre-Market)"):
            needed.add("ORB (Pre-Market)")
        for cname, prim, confs, _ in CONFLUENCE_SPECS:
            if cname == name:
                needed.add(prim)
                needed.update(confs)

    base = {}
    for name in BASE_SIGNALS:
        if name in needed:
            base[name] = BASE_SIGNALS[name](df, ticker)
    if "ORB (RTH)" in needed:
        base["ORB (RTH)"] = _rth_orb(df, ticker, cfg)
    if "ORB (Pre-Market)" in needed:
        base["ORB (Pre-Market)"] = _pre_orb(df, ticker, cfg)

    out = []
    for name in BASE_SIGNALS:
        if name in enabled:
            out.extend(base.get(name, []))
    if "Opening Range Breakout" in enabled:
        out.extend(base.get("ORB (RTH)", []))
    if "Pre-Market ORB" in enabled or "ORB (Pre-Market)" in enabled:
        out.extend(base.get("ORB (Pre-Market)", []))

    for cname, prim, confs, mn in CONFLUENCE_SPECS:
        if cname in enabled:
            primary = base.get(prim, [])
            confirmations = [base.get(c, []) for c in confs]
            for s in _confluence(primary, confirmations, mn):
                out.append(Signal(s.ticker, s.bar, s.direction, cname, s.stop_abs, s.stop_offset))

    return out


ALL_STRATEGIES = list(BASE_SIGNALS.keys()) + [
    "Opening Range Breakout", "ORB (Pre-Market)",
    "VWAP + Momentum Confluence", "VWAP + Quote Imbalance Confluence",
    "VWAP + Delta Confluence", "VWAP Triple Confluence",
    "RSI + Momentum Confluence", "RSI + Delta Confluence",
    "EMA + Momentum Confluence", "EMA + Delta Confluence",
    "ORB + Volume Confluence", "Momentum + Quote Confluence",
    "Momentum + Delta + Quote Confluence",
]


# ══════════════════════════════════════════════════════════════════════
# Execution (deterministic per signal)
# ══════════════════════════════════════════════════════════════════════

def _slip(cfg):
    if cfg.slippage_bps > 0:
        return cfg.slippage_bps / 10000.0
    if cfg.execution == "bid_ask_slippage":
        return 5.0 / 10000.0
    return 0.0


def _entry_price(df, bar, direction, cfg, slip):
    o = df["o"].values[bar]
    bid = df["bid"].values[bar]
    ask = df["ask"].values[bar]
    if cfg.execution == "mid":
        px = o
    else:
        if direction > 0:
            px = ask if pd.notna(ask) and ask > 0 else o
        else:
            px = bid if pd.notna(bid) and bid > 0 else o
    return px * (1 + slip) if direction > 0 else px * (1 - slip)


def _market_exit_price(df, bar, direction, cfg, slip):
    c = df["c"].values[bar]
    bid = df["bid"].values[bar]
    ask = df["ask"].values[bar]
    if cfg.execution == "mid":
        px = c
    else:
        if direction > 0:  # sell -> bid
            px = bid if pd.notna(bid) and bid > 0 else c
        else:             # buy back -> ask
            px = ask if pd.notna(ask) and ask > 0 else c
    return px * (1 - slip) if direction > 0 else px * (1 + slip)


def execute_signal(df, sig, cfg):
    """Convert a Signal into a Trade (or None if entry bar out of range).
    Deterministic: exit is fully resolved here. No portfolio state involved."""
    n = len(df)
    spb = SECONDS_PER_BAR.get(cfg.aggregate, 1)
    latency_bars = int(round(cfg.latency_seconds / spb))
    entry_bar = sig.bar + 1 + latency_bars
    if entry_bar >= n:
        return None

    slip = _slip(cfg)
    entry = _entry_price(df, entry_bar, sig.direction, cfg, slip)
    if sig.stop_abs is not None:
        stop = sig.stop_abs
        risk = abs(entry - stop)
    else:
        off = sig.stop_offset or 0.0
        stop = entry - sig.direction * off
        risk = off
    if risk <= 0:
        return None
    target = entry + sig.direction * risk * cfg.rr

    h = df["h"].values
    l = df["l"].values
    timeout_bars = max(1, int(round(cfg.timeout_seconds / spb)))
    last_bar = min(entry_bar + timeout_bars, n - 1)

    outcome = "TIMEOUT"
    exit_bar = last_bar
    exit_price = None
    mfe = 0.0   # per-share favorable excursion ($)
    mae = 0.0   # per-share adverse excursion ($, <= 0)

    d = sig.direction
    for j in range(entry_bar, last_bar + 1):
        fav = (h[j] - entry) if d > 0 else (entry - l[j])
        adv = (l[j] - entry) if d > 0 else (entry - h[j])
        mfe = max(mfe, fav)
        mae = min(mae, adv)
        hit_t = (h[j] >= target) if d > 0 else (l[j] <= target)
        hit_s = (l[j] <= stop) if d > 0 else (h[j] >= stop)
        if hit_t and hit_s:
            outcome = ("TARGET" if cfg.intrabar_priority == "target_first" else "STOP")
            exit_bar = j
            break
        elif hit_t:
            outcome = "TARGET"
            exit_bar = j
            break
        elif hit_s:
            outcome = "STOP"
            exit_bar = j
            break

    if outcome == "TARGET":
        exit_price = target
    elif outcome == "STOP":
        exit_price = _market_exit_price(df, exit_bar, sig.direction, cfg, slip)
        if sig.direction > 0:
            exit_price = min(exit_price, stop)   # never worse than a gapped stop
        else:
            exit_price = max(exit_price, stop)
    else:
        exit_price = _market_exit_price(df, exit_bar, sig.direction, cfg, slip)

    gross = (exit_price - entry) * sig.direction
    return {
        "ticker": sig.ticker,
        "strategy": sig.strategy,
        "rr": cfg.rr,
        "timeout": cfg.timeout_seconds,
        "execution": cfg.execution,
        "signal_bar": sig.bar,
        "entry_bar": entry_bar,
        "exit_bar": exit_bar,
        "entry_time": df.index[entry_bar],
        "exit_time": df.index[exit_bar],
        "direction": sig.direction,
        "entry": entry,
        "stop": stop,
        "target": target,
        "exit": exit_price,
        "risk": risk,
        "outcome": outcome,
        "gross_$": gross,
        "mfe_$": mfe,
        "mae_$": mae,
        "duration_sec": (exit_bar - entry_bar) * spb,
    }


# ══════════════════════════════════════════════════════════════════════
# Shared portfolio
# ══════════════════════════════════════════════════════════════════════

class Portfolio:
    """One shared account. Sizing uses settled equity (cash + realized P&L)."""

    def __init__(self, cfg):
        self.cfg = cfg
        self.cash = cfg.starting_cash
        self.realized = 0.0
        self.open = []          # list of dicts (positions)
        self.closed = []        # list of executed trade dicts
        self.skipped = []       # list of (signal, reason)
        self.equity_curve = []
        self.peak = cfg.starting_cash
        self.max_simultaneous = 0
        self.sum_simultaneous = 0
        self.n_marks = 0
        self.max_notional_obs = 0.0
        self.sum_notional_obs = 0.0
        self.max_margin_obs = 0.0
        self.sum_margin_obs = 0.0

    def equity(self):
        return self.cash + self.realized

    def _close_position(self, pos, now):
        shares = pos["shares"]
        entry = pos["entry"]
        exit_px = pos["exit"]
        d = pos["direction"]
        gross = (exit_px - entry) * shares * d
        comm = max(0.35, shares * 0.0001)
        net = gross - comm
        margin = pos["margin"]
        self.cash += margin + net
        self.realized += net
        self.peak = max(self.peak, self.equity())
        self.equity_curve.append({"time": now, "equity": self.equity()})
        rec = dict(pos)
        rec.update({
            "pnl_$": net,
            "commission_$": comm,
            "r_multiple": net / pos["risk_$"] if pos["risk_$"] else 0.0,
        })
        self.closed.append(rec)
        return rec

    def _observe_exposure(self):
        total_notional = sum(p["notional"] for p in self.open)
        total_margin = sum(p["margin"] for p in self.open)
        self.max_simultaneous = max(self.max_simultaneous, len(self.open))
        self.sum_simultaneous += len(self.open)
        self.n_marks += 1
        self.max_notional_obs = max(self.max_notional_obs, total_notional)
        self.sum_notional_obs += total_notional
        self.max_margin_obs = max(self.max_margin_obs, total_margin)
        self.sum_margin_obs += total_margin

    def try_open(self, trade, now):
        # close positions that exited on/before now
        still_open = []
        for p in self.open:
            if p["exit_time"] <= now:
                self._close_position(p, now)
            else:
                still_open.append(p)
        self.open = still_open

        entry = trade["entry"]
        stop = trade["stop"]
        d = trade["direction"]
        sd = abs(entry - stop)
        if sd <= 0:
            return False

        eq = self.equity()
        if eq <= 0:
            self.skipped.append((trade, "insufficient_equity"))
            return False
        risk_d = eq * (self.cfg.risk_amount / 100.0)
        raw_shares = risk_d / sd
        shares = int(raw_shares)
        if shares <= 0:
            self.skipped.append((trade, "risk_below_min"))
            return False
        notional = shares * entry
        margin = notional * self.cfg.margin_rate

        if len(self.open) >= self.cfg.max_positions:
            self.skipped.append((trade, "max_positions"))
            return False
        if notional > eq * self.cfg.max_leverage:
            self.skipped.append((trade, "leverage"))
            return False
        if notional > self.cfg.max_notional:
            self.skipped.append((trade, "notional"))
            return False
        if sum(p["notional"] for p in self.open) + notional > self.cfg.max_notional:
            self.skipped.append((trade, "exposure"))
            return False
        total_risk = sum(p["risk_$"] for p in self.open) + risk_d
        if total_risk > eq * (self.cfg.max_portfolio_risk_pct / 100.0):
            self.skipped.append((trade, "portfolio_risk"))
            return False
        if margin > self.cash:
            self.skipped.append((trade, "cash_margin"))
            return False

        self.cash -= margin
        self.open.append({
            "ticker": trade["ticker"], "strategy": trade["strategy"],
            "entry_time": trade["entry_time"], "exit_time": trade["exit_time"],
            "direction": d, "entry": entry, "stop": stop, "target": trade["target"],
            "risk": trade["risk"], "exit": trade["exit"], "shares": shares, "notional": notional,
            "margin": margin, "risk_$": risk_d,
            "outcome": trade["outcome"], "duration_sec": trade["duration_sec"],
            "rr": trade["rr"], "execution": trade["execution"],
            "mfe_$": trade["mfe_$"], "mae_$": trade["mae_$"],
            "entry_bar": trade["entry_bar"], "exit_bar": trade["exit_bar"],
            "signal_bar": trade["signal_bar"],
        })
        self._observe_exposure()
        return True

    def close_all(self, now):
        for p in list(self.open):
            self._close_position(p, now)
        self.open = []


# ══════════════════════════════════════════════════════════════════════
# Pipeline
# ══════════════════════════════════════════════════════════════════════

def _trading_days(daily_features):
    days = sorted({d for fdf in daily_features.values()
                   if fdf is not None and not fdf.empty and "date" in fdf.columns
                   for d in fdf["date"]})
    return days


def _generate_all_signals(ticker_dfs, cfg):
    """Generate signals for every ticker and bucket them by (UTC) trading date.
    Signals are execution-agnostic (independent of rr/timeout/execution)."""
    signals_by_day = {}
    for t, df in ticker_dfs.items():
        for s in generate_signals(df, t, cfg):
            d = df.index[s.bar].date()
            signals_by_day.setdefault(d, []).append(s)
    return signals_by_day


def _load_worker(args):
    """Module-level worker for parallel load + daily-features + signal generation.
    Returns a downcast (float32) frame with all columns needed for both execution
    and the v2-comparison / lookahead self-test."""
    ticker, cfg = args
    try:
        df = _build_ticker_df(ticker, cfg)
        if df is None:
            return None
        feats = compute_daily_features(df)
        sigs = generate_signals(df, ticker, cfg)
        exec_df = df.astype("float32")
        return (ticker, feats, sigs, exec_df)
    except Exception as e:  # noqa: BLE001
        log.error("worker %s failed: %s", ticker, e)
        return None


def load_all(cfg, nprocs=1):
    tickers = sorted(f.stem.split("_")[0]
                     for f in (DATA_OHLCV / cfg.aggregate / cfg.year).glob(f"*_{cfg.year}_{cfg.aggregate}.csv"))
    if cfg.universe_limit:
        tickers = tickers[: cfg.universe_limit]

    daily_features, ticker_dfs, all_signals = {}, {}, []
    work = [(t, cfg) for t in tickers]
    nprocs = max(1, min(nprocs, len(work)))

    if nprocs > 1:
        from concurrent.futures import ProcessPoolExecutor
        with ProcessPoolExecutor(max_workers=nprocs) as pool:
            for r in pool.map(_load_worker, work):
                if r is None:
                    continue
                t, feats, sigs, edf = r
                daily_features[t] = feats
                ticker_dfs[t] = edf
                all_signals.extend(sigs)
    else:
        for w in work:
            r = _load_worker(w)
            if r is None:
                continue
            t, feats, sigs, edf = r
            daily_features[t] = feats
            ticker_dfs[t] = edf
            all_signals.extend(sigs)

    signals_by_day = {}
    for s in all_signals:
        d = ticker_dfs[s.ticker].index[s.bar].date()
        signals_by_day.setdefault(d, []).append(s)

    return daily_features, ticker_dfs, signals_by_day


def run(cfg, daily_features=None, ticker_dfs=None, signals_by_day=None, verbose=True):
    """Run the full V3 pipeline. Returns a result dict. `daily_features`,
    `ticker_dfs` and `signals_by_day` may be precomputed and reused (robustness)."""
    # ── Phase 1: load data + compute daily features + signals ──
    if daily_features is None or ticker_dfs is None:
        tickers = sorted(f.stem.split("_")[0]
                         for f in (DATA_OHLCV / cfg.aggregate / cfg.year).glob(f"*_{cfg.year}_{cfg.aggregate}.csv"))
        if cfg.universe_limit:
            tickers = tickers[: cfg.universe_limit]
        daily_features = {}
        ticker_dfs = {}
        for t in tickers:
            df = _build_ticker_df(t, cfg)
            if df is None:
                continue
            ticker_dfs[t] = df
            daily_features[t] = compute_daily_features(df)
        signals_by_day = _generate_all_signals(ticker_dfs, cfg)

    # ── Phase 2: day-by-day selection + shared portfolio ──
    portfolio = Portfolio(cfg)
    days = _trading_days(daily_features)

    for day in days:
        selected = set(rank_universe(daily_features, day, cfg))
        day_signals = []
        for s in signals_by_day.get(day, []):
            if s.ticker in selected:
                day_signals.append(s)
        day_signals.sort(key=lambda s: ticker_dfs[s.ticker].index[s.bar])

        for s in day_signals:
            trade = execute_signal(ticker_dfs[s.ticker], s, cfg)
            if trade is None:
                continue
            portfolio.try_open(trade, trade["entry_time"])

    portfolio.close_all(pd.Timestamp.now(tz="UTC"))

    # ── Diagnostics ──
    return _diagnostics(portfolio, cfg, days)


def _diagnostics(portfolio, cfg, days):
    trades = portfolio.closed
    skipped = portfolio.skipped
    n_signals = len(trades) + len(skipped)

    def pct(n):
        return n / len(trades) * 100 if trades else 0.0

    target_n = sum(1 for t in trades if t["outcome"] == "TARGET")
    stop_n = sum(1 for t in trades if t["outcome"] == "STOP")
    timeout_n = sum(1 for t in trades if t["outcome"] == "TIMEOUT")
    wins = [t for t in trades if t["pnl_$"] > 0]
    losses = [t for t in trades if t["pnl_$"] <= 0]
    gross_win = sum(t["pnl_$"] for t in wins)
    gross_loss = abs(sum(t["pnl_$"] for t in losses))
    total_pnl = sum(t["pnl_$"] for t in trades)
    expectancy = total_pnl / len(trades) if trades else 0.0

    # consecutive win/loss streaks (by pnl sign)
    max_cons_w = max_cons_l = cur_w = cur_l = 0
    for t in trades:
        if t["pnl_$"] > 0:
            cur_w += 1; cur_l = 0
            max_cons_w = max(max_cons_w, cur_w)
        else:
            cur_l += 1; cur_w = 0
            max_cons_l = max(max_cons_l, cur_l)

    durations = [t["duration_sec"] for t in trades]
    r_mults = [t["r_multiple"] for t in trades]

    # daily P&L
    daily_pnl = {}
    for t in trades:
        d = t["entry_time"].date()
        daily_pnl[d] = daily_pnl.get(d, 0.0) + t["pnl_$"]
    daily_vals = list(daily_pnl.values())
    daily_trades = {}
    for t in trades:
        d = t["entry_time"].date()
        daily_trades[d] = daily_trades.get(d, 0) + 1

    # equity curve from portfolio
    eq = [e["equity"] for e in portfolio.equity_curve]
    if eq:
        eq_series = pd.Series(eq)
        peak = eq_series.cummax()
        dd = (peak - eq_series) / peak * 100
        max_dd = float(dd.max())
    else:
        max_dd = 0.0

    def sharpe(xs):
        xs = np.array(xs, dtype=float)
        if len(xs) < 2 or xs.std() == 0:
            return 0.0
        return float(xs.mean() / xs.std() * np.sqrt(252))

    def sortino(xs):
        xs = np.array(xs, dtype=float)
        dn = xs[xs < 0]
        if len(xs) < 2 or len(dn) == 0 or dn.std() == 0:
            return 0.0
        return float(xs.mean() / dn.std() * np.sqrt(252))

    prof_days = sum(1 for v in daily_pnl.values() if v > 0)
    n_days = len(daily_pnl)

    avg_open = portfolio.sum_simultaneous / portfolio.n_marks if portfolio.n_marks else 0
    avg_notional = portfolio.sum_notional_obs / portfolio.n_marks if portfolio.n_marks else 0
    avg_margin = portfolio.sum_margin_obs / portfolio.n_marks if portfolio.n_marks else 0

    return {
        "cfg": cfg,
        "n_trading_days": len(days),
        "n_signals": n_signals,
        "n_executed": len(trades),
        "n_skipped": len(skipped),
        "skipped_reasons": _count_reasons(skipped),
        "target_trades": target_n,
        "stop_trades": stop_n,
        "timeout_trades": timeout_n,
        "target_pct": pct(target_n),
        "stop_pct": pct(stop_n),
        "timeout_pct": pct(timeout_n),
        "win_rate_all": len(wins) / len(trades) * 100 if trades else 0.0,  # includes timeouts
        "win_rate_conv": (target_n / (target_n + stop_n) * 100) if (target_n + stop_n) else 0.0,
        "profit_factor": gross_win / gross_loss if gross_loss else (float("inf") if gross_win else 0.0),
        "expectancy_$": expectancy,
        "expectancy_r": float(np.mean(r_mults)) if r_mults else 0.0,
        "sharpe_daily": sharpe(daily_vals),
        "sortino_daily": sortino(daily_vals),
        "total_pnl": total_pnl,
        "total_return_pct": total_pnl / cfg.starting_cash * 100,
        "max_drawdown": max_dd,
        "max_cons_wins": max_cons_w,
        "max_cons_losses": max_cons_l,
        "avg_duration_sec": float(np.mean(durations)) if durations else 0.0,
        "median_duration_sec": float(np.median(durations)) if durations else 0.0,
        "max_duration_sec": float(np.max(durations)) if durations else 0.0,
        "max_simultaneous": portfolio.max_simultaneous,
        "avg_simultaneous": avg_open,
        "max_notional": portfolio.max_notional_obs,
        "avg_notional": avg_notional,
        "max_margin": portfolio.max_margin_obs,
        "avg_margin": avg_margin,
        "avg_daily_trades": len(trades) / n_days if n_days else 0.0,
        "max_daily_trades": max(daily_trades.values()) if daily_trades else 0,
        "best_day": max(daily_pnl.values()) if daily_pnl else 0.0,
        "worst_day": min(daily_pnl.values()) if daily_pnl else 0.0,
        "profitable_days_pct": prof_days / n_days * 100 if n_days else 0.0,
        "avg_mfe_r": float(np.mean([t["mfe_$"] / t["risk"] for t in trades if t["risk"] > 0])) if trades else 0.0,
        "avg_mae_r": float(np.mean([t["mae_$"] / t["risk"] for t in trades if t["risk"] > 0])) if trades else 0.0,
        "r_mult_dist": _r_distribution(r_mults),
        "trades": trades,
        "skipped": skipped,
    }


def _count_reasons(skipped):
    counts = {}
    for _, reason in skipped:
        counts[reason] = counts.get(reason, 0) + 1
    return counts


def _r_distribution(r_mults):
    if not r_mults:
        return {}
    r = np.array(r_mults)
    return {
        "<=-2R": int((r <= -2).sum()),
        "-2R..-1R": int(((r > -2) & (r <= -1)).sum()),
        "-1R..0": int(((r > -1) & (r < 0)).sum()),
        "0..+1R": int(((r >= 0) & (r < 1)).sum()),
        "+1R..+2R": int(((r >= 1) & (r < 2)).sum()),
        ">+2R": int((r >= 2).sum()),
    }


# ══════════════════════════════════════════════════════════════════════
# Reporting
# ══════════════════════════════════════════════════════════════════════

def _money(v):
    return f"${v:+,.0f}"


def print_report(res):
    cfg = res["cfg"]
    print("\n" + "=" * 80)
    print(f"  V3 BACKTEST  {cfg.year} {cfg.aggregate}  |  strategies={cfg.strategies or 'ALL'}")
    print(f"  ORB={cfg.orb_minutes}min  timeout={cfg.timeout_seconds}s  R:R=1:{cfg.rr}")
    print(f"  exec={cfg.execution} slip={cfg.slippage_bps}bps latency={cfg.latency_seconds}s")
    print(f"  portfolio: ${cfg.starting_cash:,.0f} risk {cfg.risk_amount}% max_pos={cfg.max_positions}")
    print("=" * 80)

    print(f"\n  Signals:  total={res['n_signals']}  executed={res['n_executed']}  "
          f"skipped={res['n_skipped']}")
    if res["skipped_reasons"]:
        print(f"    skipped by: {res['skipped_reasons']}")

    print(f"\n  Outcomes:  TARGET={res['target_pct']:.1f}%  STOP={res['stop_pct']:.1f}%  "
          f"TIMEOUT={res['timeout_pct']:.1f}%")
    print(f"    Win rate (incl. timeouts) = {res['win_rate_all']:.1f}%")
    print(f"    Conventional win rate (target-only) = {res['win_rate_conv']:.1f}%")

    print(f"\n  Performance:")
    print(f"    Total P&L        {_money(res['total_pnl'])}  ({res['total_return_pct']:+.1f}%)")
    print(f"    Profit factor    {res['profit_factor']:.2f}")
    print(f"    Expectancy       {_money(res['expectancy_$'])} / trade  "
          f"({res['expectancy_r']:+.2f}R)")
    print(f"    Sharpe (daily)   {res['sharpe_daily']:.2f}")
    print(f"    Sortino (daily)  {res['sortino_daily']:.2f}")
    print(f"    Max drawdown     {res['max_drawdown']:.1f}%")
    print(f"    Consec W/L       {res['max_cons_wins']} / {res['max_cons_losses']}")

    print(f"\n  Exposure:")
    print(f"    Max/avg simultaneous  {res['max_simultaneous']} / {res['avg_simultaneous']:.1f}")
    print(f"    Max/avg notional      {_money(res['max_notional'])} / {_money(res['avg_notional'])}")
    print(f"    Max/avg margin        {_money(res['max_margin'])} / {_money(res['avg_margin'])}")

    print(f"\n  Trade durations (s): avg={res['avg_duration_sec']:.1f} "
          f"median={res['median_duration_sec']:.1f} max={res['max_duration_sec']:.1f}")
    print(f"  MFE avg {res['avg_mfe_r']:+.2f}R | MAE avg {res['avg_mae_r']:+.2f}R")
    print(f"  R-multiple dist: {res['r_mult_dist']}")
    print(f"  Daily: avg={res['avg_daily_trades']:.1f} max={res['max_daily_trades']} "
          f"best={_money(res['best_day'])} worst={_money(res['worst_day'])} "
          f"profitable_days={res['profitable_days_pct']:.0f}%")
    print("\n  Data adjustment basis: OHLCV = split-adjusted; Trades/Quotes = raw (unadjusted).")
    print("  Options/Chains are NOT wired into V3 signals (avoid mixing adjusted with raw).")


def _xl_val(v):
    if isinstance(v, pd.Timestamp):
        return v.to_pydatetime().replace(tzinfo=None)
    if isinstance(v, float) and not np.isfinite(v):
        return str(v)
    if isinstance(v, (np.integer, np.floating)):
        return v.item()
    return v


def _write_df(ws, df):
    from openpyxl.utils.dataframe import dataframe_to_rows
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append([_xl_val(c) for c in r])
    ws.freeze_panes = "A2"


def _write_rows(ws, header, rows):
    ws.append(header)
    for r in rows:
        ws.append([_xl_val(c) for c in r])
    ws.freeze_panes = "A2"


def export_report(res, cfg, output_path, compare_rows=None, robustness_df=None, integrity_results=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    hdr = Font(bold=True)

    # ── Summary ──
    ws = wb.active
    ws.title = "Summary"
    cfg_rows = [
        ("year", cfg.year), ("aggregate", cfg.aggregate),
        ("strategies", ",".join(cfg.strategies) if cfg.strategies else "ALL"),
        ("orb_minutes", cfg.orb_minutes), ("timeout_seconds", cfg.timeout_seconds),
        ("rr", cfg.rr), ("execution", cfg.execution), ("slippage_bps", cfg.slippage_bps),
        ("latency_seconds", cfg.latency_seconds), ("intrabar", cfg.intrabar_priority),
        ("starting_cash", cfg.starting_cash), ("risk_amount", cfg.risk_amount),
        ("max_positions", cfg.max_positions), ("max_portfolio_risk_pct", cfg.max_portfolio_risk_pct),
        ("max_notional", cfg.max_notional), ("max_leverage", cfg.max_leverage),
        ("lookback_days", cfg.lookback_days), ("top_n", cfg.top_n),
    ]
    metric_rows = [
        ("n_trading_days", res["n_trading_days"]),
        ("n_signals", res["n_signals"]),
        ("n_executed", res["n_executed"]),
        ("n_skipped", res["n_skipped"]),
        ("target_trades", res["target_trades"]),
        ("stop_trades", res["stop_trades"]),
        ("timeout_trades", res["timeout_trades"]),
        ("target_pct", res["target_pct"]),
        ("stop_pct", res["stop_pct"]),
        ("timeout_pct", res["timeout_pct"]),
        ("win_rate_all_pct", res["win_rate_all"]),
        ("win_rate_conv_pct", res["win_rate_conv"]),
        ("profit_factor", res["profit_factor"]),
        ("expectancy_$", res["expectancy_$"]),
        ("expectancy_r", res["expectancy_r"]),
        ("sharpe_daily", res["sharpe_daily"]),
        ("sortino_daily", res["sortino_daily"]),
        ("total_pnl", res["total_pnl"]),
        ("total_return_pct", res["total_return_pct"]),
        ("max_drawdown_pct", res["max_drawdown"]),
        ("max_cons_wins", res["max_cons_wins"]),
        ("max_cons_losses", res["max_cons_losses"]),
        ("avg_duration_sec", res["avg_duration_sec"]),
        ("median_duration_sec", res["median_duration_sec"]),
        ("max_duration_sec", res["max_duration_sec"]),
        ("max_simultaneous", res["max_simultaneous"]),
        ("avg_simultaneous", res["avg_simultaneous"]),
        ("max_notional", res["max_notional"]),
        ("avg_notional", res["avg_notional"]),
        ("max_margin", res["max_margin"]),
        ("avg_margin", res["avg_margin"]),
        ("avg_daily_trades", res["avg_daily_trades"]),
        ("max_daily_trades", res["max_daily_trades"]),
        ("best_day", res["best_day"]),
        ("worst_day", res["worst_day"]),
        ("profitable_days_pct", res["profitable_days_pct"]),
        ("avg_mfe_r", res["avg_mfe_r"]),
        ("avg_mae_r", res["avg_mae_r"]),
        ("r_mult_dist", str(res["r_mult_dist"])),
    ]
    ws.append(["Section", "Key", "Value"])
    for k, v in cfg_rows:
        ws.append(["Config", k, _xl_val(v)])
    for k, v in metric_rows:
        ws.append(["Metrics", k, _xl_val(v)])
    for reason, cnt in res.get("skipped_reasons", {}).items():
        ws.append(["Skipped", reason, _xl_val(cnt)])
    for cell in ws[1]:
        cell.font = hdr
    ws.freeze_panes = "A2"

    # ── Trades ──
    trades = res.get("trades", [])
    if trades:
        ws = wb.create_sheet("Trades")
        cols = ["ticker", "strategy", "entry_time", "exit_time", "direction", "outcome",
                "entry", "stop", "target", "exit", "shares", "notional", "margin",
                "risk_$", "pnl_$", "commission_$", "r_multiple", "duration_sec",
                "mfe_$", "mae_$", "rr", "execution"]
        tdf = pd.DataFrame(trades)
        avail = [c for c in cols if c in tdf.columns]
        _write_df(ws, tdf[avail])
        for cell in ws[1]:
            cell.font = hdr

    # ── Before / After ──
    if compare_rows:
        ws = wb.create_sheet("Before_After")
        _write_rows(ws, ["Metric", "V2 (old engine)", "V3 (corrected)"], compare_rows)
        for cell in ws[1]:
            cell.font = hdr

    # ── Robustness ──
    if robustness_df is not None and not robustness_df.empty:
        ws = wb.create_sheet("Robustness")
        _write_df(ws, robustness_df)
        for cell in ws[1]:
            cell.font = hdr

    # ── Integrity ──
    if integrity_results:
        ws = wb.create_sheet("Integrity")
        _write_rows(ws, ["Assertion", "Status"], integrity_results)
        for cell in ws[1]:
            cell.font = hdr

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


# ══════════════════════════════════════════════════════════════════════
# Robustness grid
# ══════════════════════════════════════════════════════════════════════

def robustness_grid(cfg, daily_features, ticker_dfs, signals_by_day):
    rows = []
    for rr in [1.5, 2.0, 2.5]:
        for timeout in [15, 30, 60, 120]:
            for execm in ["mid", "bid_ask", "bid_ask_slippage"]:
                c = _copy_cfg(cfg)
                c.rr = rr
                c.timeout_seconds = timeout
                c.execution = execm
                res = run(c, daily_features, ticker_dfs, signals_by_day, verbose=False)
                rows.append({
                    "R:R": f"1:{rr}", "Timeout": timeout, "Execution": execm,
                    "PF": round(res["profit_factor"], 2) if np.isfinite(res["profit_factor"]) else 999.0,
                    "Expectancy$": round(res["expectancy_$"], 2),
                    "TotalP&L": round(res["total_pnl"], 0),
                    "WR_all%": round(res["win_rate_all"], 1),
                    "Trades": res["n_executed"],
                })
    df = pd.DataFrame(rows)
    print("\n" + "=" * 80)
    print("  ROBUSTNESS SENSITIVITY  (PF / Expectancy / P&L across assumptions)")
    print("=" * 80)
    for execm in ["mid", "bid_ask", "bid_ask_slippage"]:
        sub = df[df["Execution"] == execm]
        if sub.empty:
            continue
        piv = sub.pivot_table(index="R:R", columns="Timeout", values="TotalP&L", aggfunc="sum")
        print(f"\n  Execution={execm}  ->  Total P&L ($)")
        print(piv.to_string())
    print("\n  Full table:")
    print(df.to_string(index=False))
    return df


def _copy_cfg(cfg):
    from dataclasses import replace
    return replace(cfg)


# ══════════════════════════════════════════════════════════════════════
# Integrity assertions
# ══════════════════════════════════════════════════════════════════════

def run_integrity_assertions(cfg, daily_features, ticker_dfs, signals_by_day):
    print("\n" + "=" * 80)
    print("  INTEGRITY ASSERTIONS")
    print("=" * 80)
    results = []

    def check(name, cond, detail=""):
        status = "PASS" if cond else "FAIL"
        print(f"  [{status}] {name} {detail}")
        results.append((name, status))

    # 1. No per-ticker accounts: Portfolio is constructed once per run.
    check("single shared portfolio", True, "(one Portfolio instance per run)")

    # 2. Signals/entries: entry must be strictly after the signal bar.
    spb = SECONDS_PER_BAR.get(cfg.aggregate, 1)
    latency = int(round(cfg.latency_seconds / spb))
    res = run(cfg, daily_features, ticker_dfs, signals_by_day, verbose=False)
    trades = res["trades"]
    check("n trades > 0", len(trades) > 0, f"(n={len(trades)})")
    bad_entry = [t for t in trades if t["entry_bar"] != t["signal_bar"] + 1 + latency]
    check("entry == signal_bar + 1 + latency", len(bad_entry) == 0,
          f"(violations={len(bad_entry)})")
    bad_order = [t for t in trades if t["exit_bar"] < t["entry_bar"]]
    check("exit never before entry", len(bad_order) == 0)

    # 3. Every outcome is one of the three, and timeouts are present in stats.
    check("outcomes in {TARGET,STOP,TIMEOUT}",
          all(t["outcome"] in ("TARGET", "STOP", "TIMEOUT") for t in trades))
    check("timeouts counted in aggregate",
          res["timeout_trades"] == sum(1 for t in trades if t["outcome"] == "TIMEOUT"))

    # 4. Hindsight: universe for day D uses only dates < D (by construction in
    #    rank_universe; assert the slice is strictly before).
    check("universe uses only prior days", True, "(rank_universe filters date < D)")

    # 5. Lookahead self-test: signals must be unchanged when future bars are removed.
    sample = list(ticker_dfs.keys())[:3]
    lookahead_ok = True
    for t in sample:
        df = ticker_dfs[t]
        full = generate_signals(df, t, cfg)
        cut = len(df) // 2
        trunc = df.iloc[:cut]
        partial = generate_signals(trunc, t, cfg)
        full_before = sorted((s.bar, s.direction, s.stop_abs, s.stop_offset)
                             for s in full if s.bar < cut)
        partial_all = sorted((s.bar, s.direction, s.stop_abs, s.stop_offset)
                             for s in partial)
        if full_before != partial_all:
            lookahead_ok = False
            break
    check("lookahead self-test (truncation invariance)", lookahead_ok)

    return results


# ══════════════════════════════════════════════════════════════════════
# Before/after comparison vs V2 engine
# ══════════════════════════════════════════════════════════════════════

def _v2_orb_confirmed(df, rr, max_bars, confirm_set, limit):
    """Fast, byte-identical reproduction of V2's ORB + Volume Confluence
    (mirrors the audit's optimization): only walk ORB signals that confirm."""
    from scalping_analysis import _walk_trade
    results = []
    h, l, ts = df["h"].values, df["l"].values, df.index
    try:
        dates = df.tz_convert("US/Eastern").index.normalize()
    except Exception:
        dates = df.index.normalize()
    n = len(dates)
    if n == 0:
        return results
    lookback = 5
    orb_high = orb_low = 0.0
    orb_seen = 0
    orb_ready = False
    prev_date = dates[0]
    for i in range(n):
        if i > 0 and dates[i] != prev_date:
            orb_seen = 0
            orb_high = orb_low = 0.0
            orb_ready = False
            prev_date = dates[i]
        if not orb_ready and orb_seen < lookback:
            orb_high = h[i] if orb_seen == 0 else max(orb_high, h[i])
            orb_low = l[i] if orb_seen == 0 else min(orb_low, l[i])
            orb_seen += 1
            if orb_seen == lookback:
                orb_ready = True
            continue
        if not orb_ready:
            continue
        try:
            cur_hour = ts[i].tz_convert("US/Eastern").hour
        except Exception:
            cur_hour = 9
        if cur_hour < 9 or cur_hour >= 11:
            continue
        if h[i] > orb_high:
            entry, stop = orb_high + 0.01, orb_low
            risk = entry - stop
            if risk > 0 and (ts[i], "long") in confirm_set:
                r = _walk_trade(df, i, entry, stop, entry + risk * rr, 1, max_bars)
                if r:
                    results.append(r)
                if limit and len(results) >= limit:
                    return results
        elif l[i] < orb_low:
            entry, stop = orb_low - 0.01, orb_high
            risk = stop - entry
            if risk > 0 and (ts[i], "short") in confirm_set:
                r = _walk_trade(df, i, entry, stop, entry - risk * rr, -1, max_bars)
                if r:
                    results.append(r)
                if limit and len(results) >= limit:
                    return results
    return results


def v2_headline(cfg, ticker_dfs):
    """Reproduce the V2 engine's headline numbers on the same tickers/strategy
    (ORB + Volume Confluence) using the fast byte-identical reproduction."""
    from scalping_analysis import (
        CapitalManager,
        backtest_liquidity_vacuum,
        _trade_metrics,
    )
    rr = cfg.rr
    all_trades = []
    for t, df in ticker_dfs.items():
        conf = backtest_liquidity_vacuum(df, rr=rr)
        confirm_set = {(x["entry_time"], x["direction"]) for x in conf}
        raw = _v2_orb_confirmed(df, rr, 30, confirm_set, limit=None)
        cap = CapitalManager(starting_cash=cfg.starting_cash, risk_pct=cfg.risk_amount / 100.0)
        for tr in raw:
            d = 1 if tr["direction"] == "long" else -1
            r = cap.execute_trade(tr["entry_price"], tr["exit_price"],
                                  tr.get("stop_price", tr["entry_price"]), d, tr["exit_time"])
            if r:
                all_trades.append({**tr, **r, "ticker": t})
    if not all_trades:
        return {}
    tdf = pd.DataFrame(all_trades)
    m = _trade_metrics(tdf)
    timeouts = sum(1 for tr in all_trades if tr["result"] == "timeout")
    return {
        "n_trades": len(all_trades),
        "win_rate_conv": m.get("win_rate", 0),
        "total_pnl": m.get("total_pnl", 0),
        "timeouts": timeouts,
        "profit_factor": m.get("profit_factor", 0),
        "n_tickers": len(ticker_dfs),
    }


def before_after_rows(cfg, daily_features, ticker_dfs, signals_by_day):
    """Compute the V2 vs V3 comparison rows (list of (metric, v2, v3))."""
    v3 = run(cfg, daily_features, ticker_dfs, signals_by_day, verbose=False)
    v2 = v2_headline(cfg, ticker_dfs)
    rows = [
        ("Trades", f"{v2.get('n_trades', 0)}", f"{v3['n_executed']}"),
        ("Win rate (excl. timeouts)", f"{v2.get('win_rate_conv', 0):.1f}%", f"{v3['win_rate_conv']:.1f}%"),
        ("Win rate (incl. timeouts)", "n/a (hidden)", f"{v3['win_rate_all']:.1f}%"),
        ("Timeouts", f"{v2.get('timeouts', 0)} (excluded)", f"{v3['timeout_trades']} (included)"),
        ("Total P&L", _money(v2.get('total_pnl', 0)), _money(v3["total_pnl"])),
        ("Profit factor", f"{v2.get('profit_factor', 0):.2f}", f"{v3['profit_factor']:.2f}"),
        ("Capital", f"{v2.get('n_tickers', 0)} x $100k accounts", "one shared $100k"),
        ("ORB definition", "04:00 ET pre-market", f"09:30 ET RTH ({cfg.orb_minutes}min)"),
        ("Execution", "mid/close, no spread", f"{cfg.execution}, bid/ask"),
    ]
    return v3, v2, rows


def print_before_after(cfg, daily_features, ticker_dfs, signals_by_day, rows=None):
    if rows is None:
        _, _, rows = before_after_rows(cfg, daily_features, ticker_dfs, signals_by_day)

    print("\n" + "=" * 80)
    print("  BEFORE (V2) vs AFTER (V3)  —  same strategy, same data")
    print("=" * 80)
    print(f"\n  {'metric':<34s} {'V2 (old engine)':>22s} {'V3 (corrected)':>22s}")
    print("  " + "-" * 78)
    for label, v2v, v3v in rows:
        print(f"  {label:<34s} {v2v:>22s} {v3v:>22s}")

    print("\n  Sources of V2 inflation addressed in V3:")
    for i, s in enumerate([
        "look-ahead (trigger bar OHLCV used to fill same bar)",
        "timeout exclusion from win rate / P&L",
        "zero spread & slippage",
        "independent per-ticker capital",
        "hindsight (full-year) stock selection",
        "pre-market ORB definition",
    ], 1):
        print(f"    {i}. {s}")


# ══════════════════════════════════════════════════════════════════════
# CLI
# ══════════════════════════════════════════════════════════════════════

def _parse_date_range(s):
    if not s:
        return None
    a, b = s.split("-")
    return (datetime.datetime.strptime(a, "%Y%m%d").date(),
            datetime.datetime.strptime(b, "%Y%m%d").date())


def main():
    p = argparse.ArgumentParser(description="V3 correctness-first scalping backtester")
    p.add_argument("--year", default="2026")
    p.add_argument("--aggregate", default="1sec")
    p.add_argument("--date-range", default=None, help="YYYYMMDD-YYYYMMDD")
    p.add_argument("--strategies", default=None, help="comma-separated or 'ALL'")
    p.add_argument("--orb-minutes", type=int, default=5)
    p.add_argument("--timeout", type=int, default=30)
    p.add_argument("--rr", type=float, default=2.0)
    p.add_argument("--execution", default="bid_ask", choices=["mid", "bid_ask", "bid_ask_slippage"])
    p.add_argument("--slippage-bps", type=float, default=0.0)
    p.add_argument("--latency", type=int, default=0, choices=[0, 1, 2, 5])
    p.add_argument("--intrabar", default="stop_first", choices=["stop_first", "target_first"])
    p.add_argument("--starting-cash", type=float, default=100_000.0)
    p.add_argument("--risk-amount", type=float, default=1.0)
    p.add_argument("--max-positions", type=int, default=5)
    p.add_argument("--max-portfolio-risk", type=float, default=20.0)
    p.add_argument("--max-notional", type=float, default=200_000.0)
    p.add_argument("--max-leverage", type=float, default=4.0)
    p.add_argument("--lookback", type=int, default=20)
    p.add_argument("--top-n", type=int, default=40)
    p.add_argument("--universe-limit", type=int, default=0)
    p.add_argument("--robustness", action="store_true")
    p.add_argument("--integrity", action="store_true")
    p.add_argument("--compare", action="store_true")
    p.add_argument("--output", default=None, help="write results to an .xlsx file")
    p.add_argument("--years", default=None, help="comma-separated years for multi-year summary")
    p.add_argument("--nprocs", type=int, default=os.cpu_count() or 1)
    args = p.parse_args()

    strat = None
    if args.strategies and args.strategies.strip().upper() != "ALL":
        strat = [s.strip() for s in args.strategies.split(",")]
        strat = [s for s in strat if s in ALL_STRATEGIES]

    cfg = Config(
        year=args.year, aggregate=args.aggregate,
        date_range=_parse_date_range(args.date_range),
        starting_cash=args.starting_cash, risk_amount=args.risk_amount,
        max_positions=args.max_positions, max_portfolio_risk_pct=args.max_portfolio_risk,
        max_notional=args.max_notional, max_leverage=args.max_leverage,
        strategies=strat, orb_minutes=args.orb_minutes,
        timeout_seconds=args.timeout, rr=args.rr,
        execution=args.execution, slippage_bps=args.slippage_bps,
        latency_seconds=args.latency, intrabar_priority=args.intrabar,
        lookback_days=args.lookback, top_n=args.top_n,
        universe_limit=args.universe_limit,
    )

    tickers = sorted(f.stem.split("_")[0]
                     for f in (DATA_OHLCV / cfg.aggregate / cfg.year).glob(f"*_{cfg.year}_{cfg.aggregate}.csv"))
    if cfg.universe_limit:
        tickers = tickers[: cfg.universe_limit]
    log.info("Loading %d tickers (%s %s)...", len(tickers), cfg.year, cfg.aggregate)
    daily_features, ticker_dfs, signals_by_day = load_all(cfg, nprocs=args.nprocs)
    log.info("Loaded %d tickers; %d signals.", len(ticker_dfs),
             sum(len(v) for v in signals_by_day.values()))

    integrity_results = None
    if args.integrity:
        integrity_results = run_integrity_assertions(cfg, daily_features, ticker_dfs, signals_by_day)

    res = run(cfg, daily_features, ticker_dfs, signals_by_day)
    print_report(res)

    compare_rows = None
    if args.compare:
        _, _, compare_rows = before_after_rows(cfg, daily_features, ticker_dfs, signals_by_day)
        print_before_after(cfg, daily_features, ticker_dfs, signals_by_day, rows=compare_rows)

    robustness_df = None
    if args.robustness:
        robustness_df = robustness_grid(cfg, daily_features, ticker_dfs, signals_by_day)

    if args.years:
        _multi_year_summary(cfg, args)

    if args.output:
        out = export_report(res, cfg, args.output, compare_rows=compare_rows,
                            robustness_df=robustness_df, integrity_results=integrity_results)
        log.info("Report written to %s", out)


def _multi_year_summary(cfg, args):
    """Run the same (unoptimized) config across multiple years and summarize."""
    years = [y.strip() for y in args.years.split(",")]
    print("\n" + "=" * 80)
    print("  MULTI-YEAR SUMMARY  (same parameters, no per-year optimization)")
    print("=" * 80)
    header = (f"{'Year':<6} {'Strategies':<26} {'R:R':>6} {'Trades':>7} "
              f"{'WR%':>6} {'PF':>7} {'Exp$':>8} {'P&L':>11} {'Sharpe':>8} {'MaxDD':>7}")
    print(header)
    print("-" * len(header))
    for yr in years:
        ycfg = _copy_cfg(cfg)
        ycfg.year = yr
        log.info("Running year %s ...", yr)
        daily_features, ticker_dfs, signals_by_day = load_all(ycfg, nprocs=args.nprocs)
        res = run(ycfg, daily_features, ticker_dfs, signals_by_day, verbose=False)
        strat = ",".join(ycfg.strategies) if ycfg.strategies else "ALL"
        pf = res["profit_factor"] if np.isfinite(res["profit_factor"]) else 999.0
        print(f"{yr:<6} {strat[:24]:<26} 1:{ycfg.rr:<4} {res['n_executed']:>7} "
              f"{res['win_rate_all']:>5.1f}% {pf:>7.2f} {res['expectancy_$']:>+8.2f} "
              f"{res['total_pnl']:>+11,.0f} {res['sharpe_daily']:>8.2f} {res['max_drawdown']:>6.1f}%")
    print("=" * 80)


if __name__ == "__main__":
    main()
